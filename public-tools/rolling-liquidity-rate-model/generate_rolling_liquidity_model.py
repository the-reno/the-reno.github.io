from __future__ import annotations

from datetime import datetime
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, Reference

OUTPUT_FILE = "Rolling_Liquidity_Rate_Scenario_Model_Generic.xlsx"


def add_months(dt: datetime, months: int) -> datetime:
    month = dt.month + months
    year = dt.year + (month - 1) // 12
    month = (month - 1) % 12 + 1
    day = min(dt.day, monthrange(year, month)[1])
    return datetime(year, month, day)


def set_style(cell, fill=None, bold=False, font_color="000000", num_format=None):
    thin = Side(style="thin", color="B7B7B7")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if bold:
        cell.font = Font(bold=True, color=font_color)
    if num_format:
        cell.number_format = num_format
    cell.alignment = Alignment(vertical="center")


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    blue = "1F4E78"
    green = "D9EAD3"
    yellow = "FFF2CC"
    gray = "E7E6E6"

    ws.merge_cells("A1:N1")
    ws["A1"] = "Rolling Liquidity Portfolio Rate Scenario Model"
    ws["A1"].font = Font(bold=True, size=16, color=blue)

    ws.merge_cells("A2:N2")
    ws["A2"] = "Generic model. User inputs Fed scenarios and rollover curve assumptions through Dec-2026."
    ws["A2"].font = Font(italic=True, color="666666")

    # Environment
    ws.merge_cells("A4:B4")
    ws["A4"] = "Environment"
    set_style(ws["A4"], blue, True, "FFFFFF")
    env = [
        ("As Of Date", datetime(2026, 5, 26), "dd-mmm-yy"),
        ("Horizon End", datetime(2026, 12, 31), "dd-mmm-yy"),
        ("SOFR Today", 0.0355, "0.00%"),
        ("SOFR Spread", 0.0026, "0.00%"),
        ("Rollover Pass-through", 0.80, "0%"),
    ]
    for row, (label, value, fmt) in enumerate(env, 5):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        set_style(ws[f"A{row}"], green, True)
        set_style(ws[f"B{row}"], yellow, num_format=fmt)

    # Portfolio
    ws.merge_cells("D4:E4")
    ws["D4"] = "Portfolio"
    set_style(ws["D4"], blue, True, "FFFFFF")
    portfolio = [
        ("Total Portfolio", 10_000_000_000, "$#,##0"),
        ("1M Weight", 1 / 3, "0%"),
        ("2M Weight", 1 / 3, "0%"),
        ("3M Weight", 1 / 3, "0%"),
        ("Current Fixed Rate", 0.0395, "0.00%"),
        ("Proposed Floating %", 0.25, "0%"),
    ]
    for row, (label, value, fmt) in enumerate(portfolio, 5):
        ws[f"D{row}"] = label
        ws[f"E{row}"] = value
        set_style(ws[f"D{row}"], green, True)
        set_style(ws[f"E{row}"], yellow, num_format=fmt)

    # Fed scenario inputs
    ws.merge_cells("G4:I4")
    ws["G4"] = "Fed Scenario Inputs"
    set_style(ws["G4"], blue, True, "FFFFFF")
    for cell, value in zip(["G5", "H5", "I5"], ["Meeting Date", "Base Move bp", "Scenario Move bp"]):
        ws[cell] = value
        set_style(ws[cell], green, True)
    fed_dates = [datetime(2026, 6, 17), datetime(2026, 7, 29), datetime(2026, 9, 16), datetime(2026, 10, 28), datetime(2026, 12, 9)]
    for row, dt in enumerate(fed_dates, 6):
        ws[f"G{row}"] = dt
        ws[f"H{row}"] = 0
        ws[f"I{row}"] = 0
        set_style(ws[f"G{row}"], yellow, num_format="dd-mmm-yy")
        set_style(ws[f"H{row}"], yellow, num_format="0")
        set_style(ws[f"I{row}"], yellow, num_format="0")

    # Rollover rates
    ws.merge_cells("K4:M4")
    ws["K4"] = "Rollover Rate Inputs"
    set_style(ws["K4"], blue, True, "FFFFFF")
    for cell, value in zip(["K5", "L5", "M5"], ["Tenor", "Base Rate", "Scenario Adj bp"]):
        ws[cell] = value
        set_style(ws[cell], green, True)
    for row, tenor in enumerate(["1M", "2M", "3M"], 6):
        ws[f"K{row}"] = tenor
        ws[f"L{row}"] = 0.0395
        ws[f"M{row}"] = 0
        set_style(ws[f"K{row}"], green, True)
        set_style(ws[f"L{row}"], yellow, num_format="0.00%")
        set_style(ws[f"M{row}"], yellow, num_format="0")

    # Executive summary
    ws.merge_cells("A12:F12")
    ws["A12"] = "Executive Summary"
    set_style(ws["A12"], blue, True, "FFFFFF")
    headers = ["Metric", "100% Fixed - Base", "100% Fixed - Scenario", "Proposed Mix", "50/50 Mix", "100% Floating"]
    for col, val in enumerate(headers, 1):
        ws.cell(13, col, val)
        set_style(ws.cell(13, col), green, True)
    metrics = ["Income to Horizon", "Yield to Horizon", "Increment vs 100% Fixed Scenario", "Floating Allocation", "Fixed Allocation"]
    for row, metric in enumerate(metrics, 14):
        ws[f"A{row}"] = metric
        set_style(ws[f"A{row}"], green, True)

    # Schedule
    ws.merge_cells("A21:H21")
    ws["A21"] = "Rolling Investment Schedule"
    set_style(ws["A21"], blue, True, "FFFFFF")
    sched_headers = ["Bucket", "Tenor Months", "Start Date", "Maturity/Roll Date", "Period Days", "Weight", "Base Rollover Rate", "Scenario Rollover Rate"]
    for col, val in enumerate(sched_headers, 1):
        ws.cell(22, col, val)
        set_style(ws.cell(22, col), green, True)

    start = datetime(2026, 5, 26)
    horizon = datetime(2026, 12, 31)
    schedule_rows = []
    for tenor, months in [("1M", 1), ("2M", 2), ("3M", 3)]:
        s = start
        while s < horizon:
            e = min(add_months(s, months), horizon)
            schedule_rows.append((tenor, months, s, e))
            s = e
    first_sched = 23
    last_sched = first_sched + len(schedule_rows) - 1
    for row, (tenor, months, s, e) in enumerate(schedule_rows, first_sched):
        ws[f"A{row}"] = tenor
        ws[f"B{row}"] = months
        ws[f"C{row}"] = s
        ws[f"D{row}"] = e
        ws[f"E{row}"] = f"=D{row}-C{row}"
        ws[f"F{row}"] = f'=IF(A{row}="1M",$E$6,IF(A{row}="2M",$E$7,$E$8))'
        ws[f"G{row}"] = f'=IF(A{row}="1M",$L$6,IF(A{row}="2M",$L$7,$L$8))'
        ws[f"H{row}"] = f'=G{row}+((SUMIFS($I$6:$I$10,$G$6:$G$10,">="&C{row},$G$6:$G$10,"<"&D{row})/10000)*$B$9)+IF(A{row}="1M",$M$6,IF(A{row}="2M",$M$7,$M$8))/10000'
        for col in range(1, 9):
            set_style(ws.cell(row, col))
        ws[f"C{row}"].number_format = "dd-mmm-yy"
        ws[f"D{row}"].number_format = "dd-mmm-yy"
        ws[f"E{row}"].number_format = "0"
        ws[f"F{row}"].number_format = "0%"
        ws[f"G{row}"].number_format = "0.00%"
        ws[f"H{row}"].number_format = "0.00%"

    # Calculations
    ws.merge_cells("A42:F42")
    ws["A42"] = "Scenario Income Calculations"
    set_style(ws["A42"], blue, True, "FFFFFF")
    headers2 = ["Metric", "100% Fixed Base", "100% Fixed Scenario", "Proposed Mix", "50/50 Mix", "100% Floating"]
    for col, val in enumerate(headers2, 1):
        ws.cell(43, col, val)
        set_style(ws.cell(43, col), green, True)
    calc_metrics = ["Fixed Income", "Floating Income", "Total Income", "Yield to Horizon"]
    for row, metric in enumerate(calc_metrics, 44):
        ws[f"A{row}"] = metric
        set_style(ws[f"A{row}"], green, True)

    base_sum = f"SUMPRODUCT($F${first_sched}:$F${last_sched},$E${first_sched}:$E${last_sched},$G${first_sched}:$G${last_sched}/360)*$E$5"
    scen_sum = f"SUMPRODUCT($F${first_sched}:$F${last_sched},$E${first_sched}:$E${last_sched},$H${first_sched}:$H${last_sched}/360)*$E$5"
    float_sum = f"SUMPRODUCT($F${first_sched}:$F${last_sched},$E${first_sched}:$E${last_sched},($H${first_sched}:$H${last_sched}+$B$8)/360)*$E$5"
    formulas = [
        [f"={base_sum}", f"={scen_sum}", f"={scen_sum}*(1-$E$10)", f"={scen_sum}*0.5", "=0"],
        ["=0", "=0", f"={float_sum}*$E$10", f"={float_sum}*0.5", f"={float_sum}"],
        ["=B44+B45", "=C44+C45", "=D44+D45", "=E44+E45", "=F44+F45"],
        ["=B46/$E$5*360/($B$6-$B$5)", "=C46/$E$5*360/($B$6-$B$5)", "=D46/$E$5*360/($B$6-$B$5)", "=E46/$E$5*360/($B$6-$B$5)", "=F46/$E$5*360/($B$6-$B$5)"],
    ]
    for r_idx, row_formulas in enumerate(formulas, 44):
        for c_idx, formula in enumerate(row_formulas, 2):
            ws.cell(r_idx, c_idx, formula)
            set_style(ws.cell(r_idx, c_idx), num_format="$#,##0" if r_idx < 47 else "0.00%")

    # Link summary
    summary_formulas = [
        ["=B46", "=C46", "=D46", "=E46", "=F46"],
        ["=B47", "=C47", "=D47", "=E47", "=F47"],
        ["=B46-C46", "=C46-C46", "=D46-C46", "=E46-C46", "=F46-C46"],
        ["=0", "=0", "=$E$10", "=0.5", "=1"],
        ["=1-B17", "=1-C17", "=1-D17", "=1-E17", "=1-F17"],
    ]
    for r_idx, row_formulas in enumerate(summary_formulas, 14):
        for c_idx, formula in enumerate(row_formulas, 2):
            ws.cell(r_idx, c_idx, formula)
            if r_idx in [14, 16]:
                fmt = "$#,##0"
            elif r_idx == 15:
                fmt = "0.00%"
            else:
                fmt = "0%"
            set_style(ws.cell(r_idx, c_idx), num_format=fmt)

    # Sensitivity
    ws.merge_cells("H42:M42")
    ws["H42"] = "Floating Allocation Sensitivity - Scenario Path"
    set_style(ws["H42"], blue, True, "FFFFFF")
    ws["H43"] = "Floating %"
    ws["H44"] = "Total Income"
    ws["H45"] = "Increment vs 0%"
    ws["H46"] = "Yield"
    for cell in ["H43", "H44", "H45", "H46"]:
        set_style(ws[cell], green, True)
    allocations = [0, 0.25, 0.5, 0.75, 1]
    for col_idx, alloc in enumerate(allocations, 9):
        col_letter = ws.cell(43, col_idx).column_letter
        ws.cell(43, col_idx, alloc)
        ws.cell(44, col_idx, f"={scen_sum}*(1-{col_letter}$43)+{float_sum}*{col_letter}$43")
        ws.cell(45, col_idx, f"={col_letter}44-$I$44")
        ws.cell(46, col_idx, f"={col_letter}44/$E$5*360/($B$6-$B$5)")
        set_style(ws.cell(43, col_idx), yellow, num_format="0%")
        set_style(ws.cell(44, col_idx), num_format="$#,##0")
        set_style(ws.cell(45, col_idx), num_format="$#,##0")
        set_style(ws.cell(46, col_idx), num_format="0.00%")

    # Notes
    ws.merge_cells("A51:N51")
    ws["A51"] = "Notes"
    set_style(ws["A51"], blue, True, "FFFFFF")
    ws.merge_cells("A52:N55")
    ws["A52"] = "Model purpose: quantify reinvestment risk through Dec-2026 for a rolling 1M/2M/3M liquidity portfolio. User inputs Fed moves, rollover rates, and floating allocation. No client name. Simplified cash income model; not a valuation or hedge-accounting model."
    ws["A52"].font = Font(color="666666")
    ws["A52"].alignment = Alignment(wrap_text=True, vertical="top")

    # Charts
    chart1 = BarChart()
    chart1.title = "Income by Strategy"
    chart1.y_axis.title = "Income"
    data = Reference(ws, min_col=2, max_col=6, min_row=13, max_row=14)
    cats = Reference(ws, min_col=2, max_col=6, min_row=13, max_row=13)
    chart1.add_data(data, from_rows=True, titles_from_data=True)
    ws.add_chart(chart1, "H12")

    chart2 = LineChart()
    chart2.title = "Income by Floating Allocation"
    data2 = Reference(ws, min_col=9, max_col=13, min_row=43, max_row=44)
    cats2 = Reference(ws, min_col=9, max_col=13, min_row=43, max_row=43)
    chart2.add_data(data2, from_rows=True, titles_from_data=True)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "H31")

    # Formatting
    widths = {"A": 18, "B": 16, "C": 14, "D": 18, "E": 16, "F": 14, "G": 16, "H": 18, "I": 16, "J": 14, "K": 18, "L": 14, "M": 14, "N": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A13"

    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUTPUT_FILE)
    print(f"Created {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
