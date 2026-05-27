from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws.title = 'Deposit vs SOFR Swap'

ws['A1'] = 'Deposit vs SOFR Swap Analyzer'
ws['A1'].font = Font(bold=True, size=14)

# Inputs
inputs = [
    ('Notional',100000000),
    ('Deposit Start',date(2026,5,26)),
    ('Deposit End',date(2026,8,20)),
    ('Deposit Rate',0.0395),
    ('Swap Start',date(2026,6,17)),
    ('Swap End',date(2026,8,20)),
    ('SOFR Today',0.0355),
    ('SOFR Spread bp',26),
]
for i,(k,v) in enumerate(inputs,4):
    ws[f'A{i}']=k
    ws[f'B{i}']=v

# Fed table
ws['D3']='Fed Meetings'
meetings=[
    (date(2026,6,17),25),
    (date(2026,7,29),25),
    (date(2026,9,16),0),
]
for r,(d,m) in enumerate(meetings,4):
    ws[f'D{r}']=d
    ws[f'E{r}']=m

# Holidays
ws['G3']='Holiday Dates'
holidays=[date(2026,6,19),date(2026,7,3),date(2026,9,7)]
for r,h in enumerate(holidays,4):
    ws[f'G{r}']=h

# Output
labels=['Deposit Interest','Floating Received','Fixed Paid','Net Swap','Total Interest']
for r,l in enumerate(labels,20):
    ws[f'A{r}']=l

ws['B20']='=B4*B7*(B6-B5)/360'
ws['B21']='=0'
ws['B22']='=0'
ws['B23']='=0'
ws['B24']='=B20'

# Engine
headers=['Date','Business Day','Raw SOFR','Carried SOFR','Daily Factor','Compound Factor']
for c,h in enumerate(headers,10):
    ws.cell(29,c,h)

for row in range(30,151):
    if row==30:
        ws[f'J{row}']='=$B$8'
    else:
        ws[f'J{row}']=f'=IF(J{row-1}+1<$B$9,J{row-1}+1,"")'

    ws[f'K{row}']=f'=IF(J{row}="","",IF(AND(WEEKDAY(J{row},2)<=5,COUNTIF($G$4:$G$20,J{row})=0),"Y","N"))'

    ws[f'L{row}']=(
        f'=IF(J{row}="","",IF(K{row}="Y",'
        f'$B$10+SUMIFS($E$4:$E$20,$D$4:$D$20,"<"&J{row})/10000,""))'
    )

    if row==30:
        ws[f'M{row}']=f'=IF(L{row}<>"",L{row},$B$10)'
    else:
        ws[f'M{row}']=f'=IF(L{row}<>"",L{row},M{row-1})'

    ws[f'N{row}']=f'=IF(J{row}="","",1+M{row}/360)'

    if row==30:
        ws[f'O{row}']=f'=N{row}'
    else:
        ws[f'O{row}']=f'=IF(J{row}="","",O{row-1}*N{row})'

# Final coupon
ws['T8']='=LOOKUP(2,1/($O$30:$O$150<>""),$O$30:$O$150)'
ws['T9']='=T8-1 + ($B$11/10000)*(($B$9-$B$8)/360)'

swap_days='COUNTIFS($J$30:$J$150,">="&$B$8,$J$30:$J$150,"<"&$B$9)'

ws['C20']='=B20'
ws['C21']='=B4*$T$9'
ws['C22']=f'=B4*B7*{swap_days}/360'
ws['C23']='=C21-C22'
ws['C24']='=C20+C23'

wb.save('Deposit_SOFR_Swap_Analyzer_Calendar_Compounded.xlsx')
print('Workbook created')