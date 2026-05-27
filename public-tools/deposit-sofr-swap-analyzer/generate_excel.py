"""
Deposit vs SOFR Swap Analyzer

Local-friendly Python generator for a single-tab Excel workbook comparing:
1. Fixed deposit only
2. Fixed deposit plus forward-starting SOFR swap

Key methodology:
- Deposit accrues ACT/360.
- Swap fixed leg accrues ACT/360.
- Swap floating leg uses compounded SOFR in arrears.
- SOFR changes only on business days.
- On weekends and user-entered holidays, the model carries forward the previous business-day SOFR fixing.
- Daily factor = 1 + carried SOFR plus spread over 360.
- User inputs Fed scenario moves in bp; the projected SOFR path is adjusted after meeting dates.
- No PV / MTM / hedge accounting.

Dependency:
    pip install openpyxl
"""

from __future__ import annotations

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference

OUTPUT_FILE = "Deposit_SOFR_Swap_Analyzer_Calendar_Compounded.xlsx"


def style_cell(cell, fill=None, bold=False, font_color="000000", num_format=None):
    thin = Side(style="thin", color="B7B7B7")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if bold:
        cell.font = Font(bold=True, color=font_color)
    if num_format:
        cell.number_format = num_format
    cell.alignment = Alignment(vertical="center")


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Deposit vs SOFR Swap"
    ws.sheet_view.showGridLines = False

    blue = "1F4E78"
    yellow = "FFF2CC"
    green = "D9EAD3"
    gray = "E7E6E6"

    ws["A1"] = "Deposit vs SOFR Swap Analyzer"
    ws["A1"].font = Font(bold=True, size=16, color=blue)
    ws.merge_cells("A1:H1")

    ws["A3"] = "Transaction Inputs"
    style_cell(ws["A3"], blue, True, "FFFFFF")
    inputs = [
        ("Notional", 100_000_000, "$#,##0"),
        ("Deposit Start", date(2026, 5, 26), "dd-mmm-yy"),
        ("Deposit End", date(2026, 8, 20), "dd-mmm-yy"),
        ("Deposit Rate", 0.0395, "0.00%"),
        ("Swap Start", date(2026, 6, 17), "dd-mmm-yy"),
        ("Swap End", date(2026, 8, 20), "dd-mmm-yy"),
        ("SOFR Spread bp", 26, "0"),
    ]
    for row, (label, value, fmt) in enumerate(inputs, 4):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        style_cell(ws[f"A{row}"], gray, True)
        style_cell(ws[f"B{row}"], yellow, num_format=fmt)

    ws["D3"] = "Market Data"
    style_cell(ws["D3"], blue, True, "FFFFFF")
    market = [
        ("As Of Date", date(2026, 5, 26), "dd-mmm-yy"),
        ("SOFR Today", 0.0355, "0.00%"),
        ("Effective Fed Rate", 0.0433, "0.00%"),
    ]
    for row, (label, value, fmt) in enumerate(market, 4):
        ws[f"D{row}"] = label
        ws[f"E{row}"] = value
        style_cell(ws[f"D{row}"], gray, True)
        style_cell(ws[f"E{row}"], yellow, num_format=fmt)

    ws["D8"] = "Fed Scenario"
    style_cell(ws["D8"], blue, True, "FFFFFF")
    ws["D9"] = "Meeting Date"
    ws["E9"] = "Move bp"
    style_cell(ws["D9"], gray, True)
    style_cell(ws["E9"], gray, True)
    meetings = [
        (date(2026, 6, 17), 25),
        (date(2026, 7, 29), 25),
        (date(2026, 9, 16), 0),
        (date(2026, 10, 28), 0),
        (date(2026, 12, 9), 0),
    ]
    for row, (meeting, move) in enumerate(meetings, 10):
        ws[f"D{row}"] = meeting
        ws[f"E{row}"] = move
        style_cell(ws[f"D{row}"], yellow, num_format="dd-mmm-yy")
        style_cell(ws[f"E{row}"], yellow, num_format="0")

    ws["G3"] = "Holiday Calendar"
    style_cell(ws["G3"], blue, True, "FFFFFF")
    ws["G4"] = "Holiday Date"
    ws["H4"] = "Holiday Name"
    style_cell(ws["G4"], gray, True)
    style_cell(ws["H4"], gray, True)
    holiday_defaults = [
        (date(2026, 6, 19), "Juneteenth"),
        (date(2026, 7, 3), "Independence Day Observed"),
        (date(2026, 9, 7), "Labor Day"),
        (date(2026, 11, 26), "Thanksgiving"),
        (date(2026, 12, 25), "Christmas"),
    ]
    for row in range(5, 25):
        idx = row - 5
        if idx < len(holiday_defaults):
            ws[f"G{row}"] = holiday_defaults[idx][0]
            ws[f"H{row}"] = holiday_defaults[idx][1]
        style_cell(ws[f"G{row}"], yellow, num_format="dd-mmm-yy")
        style_cell(ws[f"H{row}"], yellow)

    ws["A17"] = "Output Summary"
    style_cell(ws["A17"], blue, True, "FFFFFF")
    for col, header in zip("ABCD", ["Metric", "Deposit Only", "Deposit + Swap", "Difference"]):
        ws[f"{col}18"] = header
        style_cell(ws[f"{col}18"], gray, True)

    rows = [
        (19, "Deposit Days", "=B6-B5", "=B6-B5", "=C19-B19", "0"),
        (20, "Deposit Interest", "=B4*B7*B19/360", "=B20", "=C20-B20", "$#,##0"),
        (21, "Floating Received", "=0", "=B4*($T$9-1)", "=C21-B21", "$#,##0"),
        (22, "Fixed Paid", "=0", "=B4*B7*COUNTIFS($J$30:$J$430,\">="&$B$8,$J$30:$J$430,\"<\"&$B$9)/360", "=C22-B22", "$#,##0"),
        (23, "Net Swap", "=0", "=C21-C22", "=C23-B23", "$#,##0"),
        (24, "Total Interest", "=B20", "=C20+C23", "=C24-B24", "$#,##0"),
        (25, "Maturity Value", "=B4+B24", "=B4+C24", "=C25-B25", "$#,##0"),
        (26, "Effective Yield", "=B24/B4*360/B19", "=C24/B4*360/B19", "=C26-B26", "0.00%"),
        (27, "Compounded Float Rate", "=0", "=($T$9-1)*360/COUNTIFS($J$30:$J$430,\">="&$B$8,$J$30:$J$430,\"<\"&$B$9)", "=C27-B27", "0.00%"),
    ]
    for row, metric, f1, f2, f3, fmt in rows:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = f1
        ws[f"C{row}"] = f2
        ws[f"D{row}"] = f3
        style_cell(ws[f"A{row}"], gray, True)
        for col in "BCD":
            style_cell(ws[f"{col}{row}"], green, num_format=fmt)

    ws["S3"] = "Scenario"
    ws["T3"] = "Total Interest"
    ws["S4"] = "Deposit Only"
    ws["T4"] = "=B24"
    ws["S5"] = "Deposit + Swap"
    ws["T5"] = "=C24"
    for cell in ["S3", "T3", "S4", "T4", "S5", "T5"]:
        style_cell(ws[cell], gray if cell in ["S3", "T3"] else None, bold=cell in ["S3", "T3"], num_format="$#,##0" if cell in ["T4", "T5"] else None)

    chart = BarChart()
    chart.title = "Total Interest Comparison"
    chart.y_axis.title = "Interest"
    chart.x_axis.title = "Scenario"
    data = Reference(ws, min_col=20, min_row=3, max_row=5)
    cats = Reference(ws, min_col=19, min_row=4, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, "F17")

    ws["J28"] = "SOFR Daily Compounding Engine"
    style_cell(ws["J28"], blue, True, "FFFFFF")
    engine_headers = [
        "Date", "Bus Day", "Holiday", "Raw SOFR", "Carried SOFR", "SOFR+Spread", "Daily Factor", "Cumulative Factor"
    ]
    for col_idx, header in enumerate(engine_headers, 10):
        cell = ws.cell(29, col_idx)
        cell.value = header
        style_cell(cell, gray, True)

    for row in range(30, 431):
        if row == 30:
            ws[f"J{row}"] = "=$B$8"
        else:
            ws[f"J{row}"] = f"=IF(J{row-1}+1<$B$9,J{row-1}+1,\"\")"

        ws[f"K{row}"] = f"=IF(J{row}=\"\",\"\",IF(AND(WEEKDAY(J{row},2)<=5,COUNTIF($G$5:$G$24,J{row})=0),\"Y\",\"N\"))"
        ws[f"L{row}"] = f"=IF(J{row}=\"\",\"\",IFERROR(VLOOKUP(J{row},$G$5:$H$24,2,FALSE),\"\"))"

        # Raw SOFR changes only on business days.
        ws[f"M{row}"] = (
            f"=IF(J{row}=\"\",\"\",IF(K{row}=\"Y\",$E$5+(IF(AND($D$10<>\"\",J{row}>$D$10),$E$10,0)+"
            f"IF(AND($D$11<>\"\",J{row}>$D$11),$E$11,0)+IF(AND($D$12<>\"\",J{row}>$D$12),$E$12,0)+"
            f"IF(AND($D$13<>\"\",J{row}>$D$13),$E$13,0)+IF(AND($D$14<>\"\",J{row}>$D$14),$E$14,0))/10000,\"\"))"
        )

        # Carried SOFR repeats the previous business-day fixing on weekends and holidays.
        if row == 30:
            ws[f"N{row}"] = f"=IF(J{row}=\"\",\"\",IF(M{row}<>\"\",M{row},$E$5))"
        else:
            ws[f"N{row}"] = f"=IF(J{row}=\"\",\"\",IF(M{row}<>\"\",M{row},N{row-1}))"

        ws[f"O{row}"] = f"=IF(J{row}=\"\",\"\",N{row}+$B$10/10000)"
        ws[f"P{row}"] = f"=IF(J{row}=\"\",\"\",1+O{row}/360)"
        if row == 30:
            ws[f"Q{row}"] = f"=IF(J{row}=\"\",\"\",P{row})"
        else:
            ws[f"Q{row}"] = f"=IF(J{row}=\"\",\"\",Q{row-1}*P{row})"

        for col_idx in range(10, 18):
            style_cell(ws.cell(row, col_idx))
        ws[f"J{row}"].number_format = "dd-mmm-yy"
        ws[f"M{row}"].number_format = "0.00%"
        ws[f"N{row}"].number_format = "0.00%"
        ws[f"O{row}"].number_format = "0.00%"
        ws[f"P{row}"].number_format = "0.000000000"
        ws[f"Q{row}"].number_format = "0.000000000"

    ws["S8"] = "Final Compound Factor"
    ws["T8"] = "=LOOKUP(2,1/($Q$30:$Q$430<>\"\"),$Q$30:$Q$430)"
    ws["S9"] = "Floating Coupon Factor"
    ws["T9"] = "=T8"
    for cell in ["S8", "S9"]:
        style_cell(ws[cell], gray, True)
    for cell in ["T8", "T9"]:
        style_cell(ws[cell], green, num_format="0.000000000")

    ws["A31"] = "Methodology"
    style_cell(ws["A31"], blue, True, "FFFFFF")
    ws.merge_cells("A32:H35")
    ws["A32"] = (
        "SOFR floating leg uses one row per calendar day. SOFR is updated only on business days. "
        "On weekends and holidays, the prior business-day SOFR fixing is carried forward. "
        "Each daily factor = 1 + (carried SOFR + spread) / 360. "
        "The coupon factor is the product of all daily factors over the swap period."
    )
    ws["A32"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A32"].font = Font(color="666666")

    widths = {
        "A": 22, "B": 18, "C": 18, "D": 18, "E": 14,
        "F": 18, "G": 18, "H": 22, "J": 14, "K": 10,
        "L": 22, "M": 14, "N": 14, "O": 14, "P": 16, "Q": 18,
        "S": 22, "T": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A18"
    return wb


if __name__ == "__main__":
    workbook = build_workbook()
    workbook.save(OUTPUT_FILE)
    print(f"Created {OUTPUT_FILE}")
