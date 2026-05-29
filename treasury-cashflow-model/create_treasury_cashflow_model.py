from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_FILE = "treasury_cashflow_model.xlsx"


def add_months(d, months):
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    days_in_month = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(d.day, days_in_month[month - 1])
    return date(year, month, day)


def is_business_day(d, holidays):
    return d.weekday() < 5 and d not in holidays


def adjust_business_day(d, holidays, convention="Following"):
    convention = convention.lower()

    if is_business_day(d, holidays):
        return d

    if convention == "preceding":
        while not is_business_day(d, holidays):
            d -= timedelta(days=1)
        return d

    if convention == "modified following":
        original_month = d.month
        adjusted = d
        while not is_business_day(adjusted, holidays):
            adjusted += timedelta(days=1)
        if adjusted.month != original_month:
            adjusted = d
            while not is_business_day(adjusted, holidays):
                adjusted -= timedelta(days=1)
        return adjusted

    # default: Following
    while not is_business_day(d, holidays):
        d += timedelta(days=1)
    return d


def year_fraction_30_360_us(start, end):
    d1 = min(start.day, 30)
    d2 = end.day
    if d1 == 30 and d2 == 31:
        d2 = 30
    return ((end.year - start.year) * 360 + (end.month - start.month) * 30 + (d2 - d1)) / 360


def year_fraction_act_360(start, end):
    return (end - start).days / 360


def year_fraction_act_365(start, end):
    return (end - start).days / 365


def year_fraction(start, end, day_count):
    dc = str(day_count).upper().replace(" ", "")
    if dc in ["30/360", "30/360US", "BOND"]:
        return year_fraction_30_360_us(start, end)
    if dc in ["ACT/360", "ACTUAL/360"]:
        return year_fraction_act_360(start, end)
    if dc in ["ACT/365", "ACTUAL/365"]:
        return year_fraction_act_365(start, end)
    return year_fraction_30_360_us(start, end)


def generate_coupon_cashflows(issue_date, first_coupon, maturity_date, outstanding, coupon, freq_months, day_count, holidays, business_day_convention):
    flows = {}
    prev_coupon = issue_date
    scheduled_coupon = first_coupon

    while scheduled_coupon <= maturity_date:
        accrual_start = prev_coupon
        accrual_end = scheduled_coupon
        yf = year_fraction(accrual_start, accrual_end, day_count)
        coupon_amount = outstanding * coupon * yf

        scheduled_payment = scheduled_coupon
        settlement_date = adjust_business_day(scheduled_payment, holidays, business_day_convention)

        amount = coupon_amount
        if scheduled_coupon == maturity_date:
            amount += outstanding

        flows[settlement_date] = flows.get(settlement_date, 0) + amount

        prev_coupon = scheduled_coupon
        scheduled_coupon = add_months(scheduled_coupon, freq_months)

    return flows


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow"

    navy = PatternFill("solid", fgColor="17365D")
    light_blue = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["B3"] = "Start Date"
    ws["C3"] = date(2026, 5, 29)
    ws["B4"] = "End Date"
    ws["C4"] = date(2031, 12, 31)
    ws["B3"].font = bold_font
    ws["B4"].font = bold_font

    input_rows = {
        "Name": 8,
        "Identifier / CUSIP": 9,
        "Issue Date": 10,
        "Maturity Date": 11,
        "Outstanding": 12,
        "Coupon": 13,
        "Freq. Months": 14,
        "First Coupon Date": 15,
        "Day Count": 16,
        "Business Day Convention": 17,
    }

    for label, row in input_rows.items():
        ws.cell(row, 2, label)
        ws.cell(row, 2).font = bold_font
        ws.cell(row, 2).fill = light_blue

    instruments = [
        {
            "name": "Bond A",
            "id": "ID-A",
            "issue": date(2021, 11, 19),
            "maturity": date(2026, 11, 19),
            "outstanding": 500_000_000,
            "coupon": 0.0350,
            "freq_months": 6,
            "first_coupon": date(2022, 5, 19),
            "day_count": "30/360",
            "business_day_convention": "Following",
        },
        {
            "name": "Bond B",
            "id": "ID-B",
            "issue": date(2022, 9, 15),
            "maturity": date(2027, 9, 15),
            "outstanding": 475_000_000,
            "coupon": 0.0400,
            "freq_months": 6,
            "first_coupon": date(2023, 3, 15),
            "day_count": "30/360",
            "business_day_convention": "Following",
        },
        {
            "name": "Bond C",
            "id": "ID-C",
            "issue": date(2024, 5, 14),
            "maturity": date(2034, 5, 14),
            "outstanding": 300_000_000,
            "coupon": 0.0550,
            "freq_months": 6,
            "first_coupon": date(2024, 11, 14),
            "day_count": "30/360",
            "business_day_convention": "Following",
        },
    ]

    start_col = 3
    for i, inst in enumerate(instruments):
        col = start_col + i
        ws.cell(input_rows["Name"], col, inst["name"])
        ws.cell(input_rows["Identifier / CUSIP"], col, inst["id"])
        ws.cell(input_rows["Issue Date"], col, inst["issue"])
        ws.cell(input_rows["Maturity Date"], col, inst["maturity"])
        ws.cell(input_rows["Outstanding"], col, inst["outstanding"])
        ws.cell(input_rows["Coupon"], col, inst["coupon"])
        ws.cell(input_rows["Freq. Months"], col, inst["freq_months"])
        ws.cell(input_rows["First Coupon Date"], col, inst["first_coupon"])
        ws.cell(input_rows["Day Count"], col, inst["day_count"])
        ws.cell(input_rows["Business Day Convention"], col, inst["business_day_convention"])

    holiday_col = 7
    ws.cell(21, holiday_col, "Holiday Date")
    ws.cell(21, holiday_col).font = bold_font

    sample_holidays = [
        date(2026, 1, 1),
        date(2026, 1, 19),
        date(2026, 2, 16),
        date(2026, 5, 25),
        date(2026, 7, 3),
        date(2026, 9, 7),
        date(2026, 11, 26),
        date(2026, 12, 25),
    ]
    for r, h in enumerate(sample_holidays, start=22):
        ws.cell(r, holiday_col, h)

    holidays = set(sample_holidays)
    instrument_flows = []
    for inst in instruments:
        flows = generate_coupon_cashflows(
            issue_date=inst["issue"],
            first_coupon=inst["first_coupon"],
            maturity_date=inst["maturity"],
            outstanding=inst["outstanding"],
            coupon=inst["coupon"],
            freq_months=inst["freq_months"],
            day_count=inst["day_count"],
            holidays=holidays,
            business_day_convention=inst["business_day_convention"],
        )
        instrument_flows.append(flows)

    cf_header_row = 21
    headers = ["Date", "Day Type"] + [inst["name"] for inst in instruments] + ["Total"]
    for c, h in enumerate(headers, start=1):
        ws.cell(cf_header_row, c, h)
        ws.cell(cf_header_row, c).font = white_font
        ws.cell(cf_header_row, c).fill = navy
        ws.cell(cf_header_row, c).alignment = Alignment(horizontal="center")

    start_date = ws["C3"].value
    end_date = ws["C4"].value
    current = start_date
    row = cf_header_row + 1

    while current <= end_date:
        ws.cell(row, 1, current)
        ws.cell(row, 2, f'=IF(COUNTIF($G$22:$G$200,A{row})>0,"HOL",IF(WEEKDAY(A{row},2)>5,"WE","BD"))')

        for i, flows in enumerate(instrument_flows):
            ws.cell(row, 3 + i, flows.get(current, 0))

        total_col = 3 + len(instruments)
        ws.cell(row, total_col, f"=SUM(C{row}:{get_column_letter(total_col-1)}{row})")
        current += timedelta(days=1)
        row += 1

    widths = {"A": 16, "B": 26, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for r in range(1, row):
        for c in range(1, holiday_col + 1):
            ws.cell(r, c).border = border

    for r in range(3, 5):
        ws.cell(r, 3).number_format = "dd-mmm-yyyy"

    for col in range(3, 6):
        ws.cell(10, col).number_format = "dd-mmm-yyyy"
        ws.cell(11, col).number_format = "dd-mmm-yyyy"
        ws.cell(12, col).number_format = "#,##0"
        ws.cell(13, col).number_format = "0.00%"
        ws.cell(15, col).number_format = "dd-mmm-yyyy"

    for r in range(cf_header_row + 1, row):
        ws.cell(r, 1).number_format = "dd-mmm-yyyy"
        for c in range(3, 3 + len(instruments) + 1):
            ws.cell(r, c).number_format = '#,##0.00;[Red](#,##0.00);-'

    for r in range(22, 200):
        ws.cell(r, holiday_col).number_format = "dd-mmm-yyyy"

    ws.freeze_panes = "A22"

    last_col = get_column_letter(3 + len(instruments))
    tab = Table(displayName="CashFlowTable", ref=f"A21:{last_col}{row-1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    build_workbook()
    print(f"Created: {OUTPUT_FILE}")
