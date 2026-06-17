"""build_engine_v2.py  ->  NII_Engine_v2.xlsx
Local builder for the v2 engine example workbook (object layer:
HolidayTable, FedScenario, RatesCurve, GET, ACCRUE, SWAP).

Ships formula-free; import the .bas modules from ./bas, then paste the
formulas shown in column E of the MODEL sheet.

Run:  python3 build_engine_v2.py        (needs: pip install openpyxl)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.workbook.properties import CalcProperties
from datetime import date

A="Arial"; NAVY="1F3864"
BLUE=Font(name=A,color="0000FF"); BLK=Font(name=A,size=10); SUB=Font(name=A,bold=True,size=10)
YEL=PatternFill("solid",start_color="FFF2CC"); MONO=Font(name="Consolas",size=9,color="7F4F00")
thin=Side(style="thin",color="BFBFBF"); BOX=Border(left=thin,right=thin,top=thin,bottom=thin)

def put(ws,r,c,v,font=BLK,fill=None,nf=None,box=False):
    x=ws.cell(row=r,column=c,value=v); x.font=font
    if fill:x.fill=fill
    if nf:x.number_format=nf
    if box:x.border=BOX
    return x

def build(path="NII_Engine_v2.xlsx"):
    wb=Workbook()

    # README
    ws=wb.active; ws.title="README"; ws.sheet_view.showGridLines=False
    L=["NII ENGINE v2 — curve as the named object","",
    "Four functions:",
    "  RatesCurve(name,start,end,sofr,scenRange,holRange)  build the curve",
    "  GET(curveCell)            spill the curve: provenance + daily strip",
    "  ACCRUE(start,end,amount,type,curveCell)             interest $mm",
    "  SWAP(start,end,notional,fixed,curveCell,leg)        FIXED|FLOAT|NET","",
    "The curve is the one named object. It reads the scenario and holiday",
    "ranges directly, records which ranges built it (provenance), and stores",
    "the daily strip: date | rate% | dayFactor | accumFactor.","",
    "SETUP",
    "1. Alt+F11 > File > Import File... import the 3 .bas from bas/",
    "   (mRegistry, cRatesCurve, mEngine).",
    "2. Save as .xlsm. Press Ctrl+Alt+F9.",
    "3. MODEL sheet: paste each column-E formula into the yellow cell.",
    "   Build the curve (B6) first; the rest reference it.",
    "4. Results should match Expected; GET($B$6) spills the strip.","",
    "RULE: scenRange, holRange and curveCell are passed as RANGES/CELLS,",
    "never as typed text, so editing inputs cascades automatically."]
    for i,t in enumerate(L,1):
        ws.cell(row=i,column=1,value=t).font=(Font(name=A,bold=True,size=15,color=NAVY) if i==1 else
                                              (MONO if t.startswith("  ") else BLK))
    ws.column_dimensions["A"].width=78

    # INPUTS
    ws=wb.create_sheet("INPUTS"); ws.sheet_view.showGridLines=False
    put(ws,1,1,"INPUTS",Font(name=A,bold=True,size=12,color=NAVY))
    put(ws,3,1,"Holidays (date | name)",SUB)
    hol=[(date(2026,1,1),"New Year"),(date(2026,1,19),"MLK Day"),(date(2026,7,3),"Independence (obs)"),
         (date(2026,9,7),"Labor Day"),(date(2026,11,26),"Thanksgiving"),(date(2026,12,25),"Christmas")]
    for i,(d,n) in enumerate(hol,4):
        put(ws,i,1,d,BLUE,nf="YYYY-MM-DD",box=True); put(ws,i,2,n,BLUE,box=True)
    put(ws,3,4,"FOMC moves (date | bps)",SUB)
    fomc=[(date(2026,6,17),-25),(date(2026,7,29),-25),(date(2026,9,16),-25),(date(2026,10,28),0)]
    for i,(d,m) in enumerate(fomc,4):
        put(ws,i,4,d,BLUE,nf="YYYY-MM-DD",box=True); put(ws,i,5,m,BLUE,nf='0;-0;"-"',box=True)
    put(ws,11,1,"Holidays = A4:B9   ·   FOMC = D4:E7",Font(name=A,size=9,color="595959"))
    for c,w in [("A",13),("B",20),("D",13),("E",10)]: ws.column_dimensions[c].width=w

    # MODEL
    ws=wb.create_sheet("MODEL"); ws.sheet_view.showGridLines=False
    put(ws,1,1,"MODEL — paste each column-E formula into the yellow cell",Font(name=A,bold=True,size=12,color=NAVY))
    put(ws,3,1,"1 · BUILD THE CURVE (reads the input ranges directly)",SUB); put(ws,3,5,"PASTE THIS",SUB)
    put(ws,6,1,"Curve"); put(ws,6,2,None,fill=YEL,box=True)
    put(ws,6,5,'=BuildCurve("curve1",DATE(2026,6,15),DATE(2030,12,31),3.80,INPUTS!D4:E7,INPUTS!A4:B9)',MONO)

    put(ws,8,1,"2 · ACCRUE",SUB); put(ws,8,3,"Expected",SUB); put(ws,8,4,"Check",SUB); put(ws,8,5,"PASTE THIS",SUB)
    cases=[("Period simple (Jul-Oct)",0.853750,'=Accrue(DATE(2026,7,1),DATE(2026,10,1),100,"SIMPLE",$B$6)'),
    ("Period compound (Jul-Oct)",0.857325,'=Accrue(DATE(2026,7,1),DATE(2026,10,1),100,"COMPOUND",$B$6)')]
    r=9
    for nm,exp,f in cases:
        put(ws,r,1,nm); x=put(ws,r,2,None,fill=YEL,box=True); x.number_format="0.000000"
        put(ws,r,3,exp,nf="0.000000"); put(ws,r,4,'=IF(ABS(B%d-C%d)<0.000001,"OK","check")'%(r,r),SUB)
        put(ws,r,5,f,MONO); r+=1

    put(ws,r+1,1,"3 · SWAP — $375mm @3.15%, July, receive-fixed",SUB); r+=2
    for nm,exp,f in [("FIXED leg",1.017187,'=SwapLeg(DATE(2026,7,1),DATE(2026,8,1),375,3.15,$B$6,"FIXED")'),
    ("FLOAT leg",1.142773,'=SwapLeg(DATE(2026,7,1),DATE(2026,8,1),375,3.15,$B$6,"FLOAT")'),
    ("NET",-0.125585,'=SwapLeg(DATE(2026,7,1),DATE(2026,8,1),375,3.15,$B$6,"NET")')]:
        put(ws,r,1,nm); x=put(ws,r,2,None,fill=YEL,box=True); x.number_format="0.000000"
        put(ws,r,3,exp,nf="0.000000"); put(ws,r,4,'=IF(ABS(B%d-C%d)<0.000001,"OK","check")'%(r,r),SUB)
        put(ws,r,5,f,MONO); r+=1

    put(ws,r+1,1,"4 · INSPECT — GET spills provenance + the daily strip",SUB); r+=2
    put(ws,r,1,"Curve table"); put(ws,r,5,"=CurveName($B$6)",MONO); r+=1
    put(ws,r+1,1,"Spills: header (curve/SOFR), then date | rate% | dayFactor | accumFactor.",Font(name=A,size=9,color="595959"))
    put(ws,r+2,1,"Give it empty room below and to the right.",Font(name=A,size=9,color="595959"))
    for c,w in [("A",28),("B",13),("C",13),("D",8),("E",64)]: ws.column_dimensions[c].width=w

    wb.calculation=CalcProperties(fullCalcOnLoad=True)
    wb.save(path); print("saved:",path)

if __name__=="__main__":
    build()
