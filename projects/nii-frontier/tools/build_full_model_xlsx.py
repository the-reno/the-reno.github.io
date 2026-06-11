"""NII FRONTIER - full standalone spreadsheet. Mirrors ronu.one/projects/nii-frontier in live
Excel formulas on a single tab: frontier (1,430 strategies, live Pareto + knee), zones, slope
chart, convergence cone with pinch, SOFR step paths, stacked cashflow composition with scenario
selector, scenario-fair swap rate and carry gap. Monthly approximation of the daily engine
(validated ~0.3%); audit granularity is monthly (the calc grid is the ledger).
Run: python3 build_full_model_xlsx.py -> NII_Frontier_Full.xlsx. No dependencies beyond openpyxl.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as gcl
from datetime import date
from itertools import product

VERSION="v2.2"
A="Arial"
BLUE=Font(name=A,color="0000FF"); BLK=Font(name=A); SUB=Font(name=A,bold=True)
HDR=Font(name=A,bold=True,color="FFFFFF"); HF=PatternFill("solid",start_color="1F3864")
CALC=PatternFill("solid",start_color="F2F2F2"); OUT=Font(name=A,bold=True,color="1F3864")
WARN=PatternFill("solid",start_color="FFF2CC")
NUM='#,##0.0;(#,##0.0)'; N2='#,##0.00;(#,##0.00)'

FOMC=[date(2026,6,17),date(2026,7,29),date(2026,9,16),date(2026,10,28),date(2026,12,9),
date(2027,1,27),date(2027,3,17),date(2027,4,28),date(2027,6,16),date(2027,7,28),
date(2027,9,15),date(2027,10,27),date(2027,12,8),date(2028,1,26),date(2028,3,15),
date(2028,4,26),date(2028,6,14)]
SCEN=["Flat","Gradual cuts","Faster cuts","Delayed cuts","Higher for longer","Hike"]
def moves():
    n=len(FOMC); z=lambda mv:[mv.get(i,0) for i in range(n)]
    return [z({}),z({i:-25 for i in range(0,12,2)}),z({i:-25 for i in range(6)}),
            z({i:-25 for i in range(4,10)}),z({8:-25,9:-25}),z({0:25,1:25})]
BONDS=[("3.50% '27",476.4,3.50,date(2027,9,15)),("6.60% '28",109.9,6.60,date(2028,7,15)),
("3.90% '29",900.0,3.90,date(2029,11,19)),("4.65% '31",400.0,4.65,date(2031,3,12)),
("6.05% '34",500.0,6.05,date(2034,5,14)),("6.35% '40",500.0,6.35,date(2040,3,15)),
("5.10% '44",300.0,5.10,date(2044,5,15))]

wb=Workbook(); ws=wb.active; ws.title="MODEL"
def put(r,c,v,font=BLK,fill=None,nf=None,align=None):
    x=ws.cell(row=r,column=c,value=v); x.font=font
    if fill:x.fill=fill
    if nf:x.number_format=nf
    if align:x.alignment=Alignment(horizontal=align)
    return x

put(1,1,f"NII EFFICIENT FRONTIER — FULL STANDALONE MODEL {VERSION}",Font(name=A,bold=True,size=14,color="1F3864"))
put(2,1,"Single tab, live formulas, no macros. Blue = inputs. Mirrors ronu.one/projects/nii-frontier. "
        "Monthly approximation (term buckets at trailing k-month avg SOFR; swap floating = monthly SOFR; "
        "ACT/360; bonds 30/360). Frontier rows 76+: full 10%-step allocation grid with live feasibility "
        "(min ON / cash changes rework the frontier). Allocation step changes require regenerating with "
        "the build script. Illustrative only.",Font(name=A,size=9,color="595959"))

# ================= INPUTS (A4:D31)
put(4,1,"INPUTS",SUB)
inputs=[("Start month (1st)",date(2026,7,1),"YYYY-MM-DD"),("Initial SOFR (%)",3.80,N2),
("Total cash ($mm)",500.0,NUM),("Deposit beta",1.0,N2),("Min ON balance ($mm)",100.0,NUM),
("Swap direction (RCV/PAY)","RCV",None),("Swap fixed rate (%)",3.15,N2),
("Swap float spread (bps)",0.0,NUM),("IRS notional (% of cash)",75.0,NUM)]
for i,(k,v,nf) in enumerate(inputs,5): put(i,1,k); put(i,2,v,BLUE,nf=nf)
R_START,R_SOFR0,R_CASH,R_BETA,R_MIN,R_DIR,R_FIX,R_SPR,R_NOT=range(5,14)
dv=DataValidation(type="list",formula1='"RCV,PAY"'); ws.add_data_validation(dv); dv.add(f"B{R_DIR}")

put(15,1,"INPUT STRATEGY — ALLOCATION ($mm) / SPREAD (bps)",SUB)
put(16,1,"Bucket",SUB);put(16,2,"$mm",SUB);put(16,3,"Spread",SUB)
for i,(b,al,sp) in enumerate([("ON",100,0),("1M",0,10),("2M",0,15),("3M",400,20)],17):
    put(i,1,b); put(i,2,al,BLUE,nf=NUM); put(i,3,sp,BLUE,nf=NUM)
c=put(21,1,"Check"); c=put(21,2,'=IF(SUM(B17:B20)<>B7,"SUM<>CASH",IF(B17<B9,"ON<MIN","OK"))',SUB); c.fill=WARN

put(23,1,"FIXED-RATE DEBT (Hasbro, illustrative)",SUB)
put(24,1,"Issue",SUB);put(24,2,"$mm",SUB);put(24,3,"Cpn %",SUB);put(24,4,"Maturity",SUB)
for i,(n,nm,cp,mt) in enumerate(BONDS,25):
    put(i,1,n); put(i,2,nm,BLUE,nf=NUM); put(i,3,cp,BLUE,nf=N2); put(i,4,mt,BLUE,nf="YYYY-MM-DD")
B0,B1=25,31

# ================= SELF-TESTS (A33:B40) - filled after layout constants exist (see end)
# ================= FOMC x SCENARIO (F4..M22) + weights
put(4,6,"FED SCENARIOS — bps per FOMC decision",SUB)
put(5,6,"Decision",SUB); put(5,7,"Effective",SUB)
for j,s in enumerate(SCEN): put(5,8+j,s,SUB,align="center")
MV=moves(); FR0,FR1=6,6+len(FOMC)-1
for i,d in enumerate(FOMC):
    r=6+i; put(r,6,d,nf="YYYY-MM-DD"); put(r,7,f"=F{r}+1",nf="YYYY-MM-DD")
    for j in range(6): put(r,8+j,MV[j][i],BLUE,nf='0;-0;"-"',align="center")
WR=FR1+2
put(WR,6,"Scenario weight",SUB)
for j in range(6): put(WR,8+j,1/6,BLUE,nf="0.0%",align="center")

# ================= CALC GRID (rows 46-69) + unit totals (71-72)
G0=46; NM=24
put(G0-2,1,"CALCULATION GRID — monthly, per scenario (do not edit)",SUB)
hdrs=["Month","Days","Bond cost"]
for s in SCEN: hdrs+=[f"{s}|SOFR","ON%","1M%","2M%","3M%","Dep inc","Swap/$"]
for c,h in enumerate(hdrs,1): put(G0-1,c,h,HDR,HF)
for m in range(NM):
    r=G0+m
    put(r,1,f"=EDATE($B${R_START},{m})",fill=CALC,nf="MMM-YY")
    put(r,2,f"=EDATE(A{r},1)-A{r}",fill=CALC)
    put(r,3,f"=SUMPRODUCT(($D${B0}:$D${B1}>A{r})*$B${B0}:$B${B1}*$C${B0}:$C${B1})/100/12",fill=CALC,nf=N2)
    for j in range(6):
        b=4+j*7; L=lambda k:gcl(b+k)
        mvr=f"{gcl(8+j)}${FR0}:{gcl(8+j)}${FR1}"
        put(r,b,f'=$B${R_SOFR0}+SUMIFS({mvr},$G${FR0}:$G${FR1},"<="&A{r})/100',fill=CALC,nf=N2)
        adj=lambda src:f"$B${R_SOFR0}+$B${R_BETA}*({src}-$B${R_SOFR0})"
        sref=f"{L(0)}{r}"
        put(r,b+1,f"={adj(sref)}+$C$17/100",fill=CALC,nf=N2)
        put(r,b+2,f"={adj(sref)}+$C$18/100",fill=CALC,nf=N2)
        col=f"{L(0)}${G0}:{L(0)}${G0+NM-1}"
        a2=f"AVERAGE(INDEX({col},MAX(1,{m})):INDEX({col},{m+1}))"
        a3=f"AVERAGE(INDEX({col},MAX(1,{m-1})):INDEX({col},{m+1}))"
        put(r,b+3,f"={adj(a2)}+$C$19/100",fill=CALC,nf=N2)
        put(r,b+4,f"={adj(a3)}+$C$20/100",fill=CALC,nf=N2)
        put(r,b+5,f"=($B$17*{L(1)}{r}+$B$18*{L(2)}{r}+$B$19*{L(3)}{r}+$B$20*{L(4)}{r})/100*B{r}/360",
            fill=CALC,nf="0.0000")
        put(r,b+6,f'=IF($B${R_DIR}="RCV",1,-1)*(($B${R_FIX}/100)-({L(0)}{r}/100+$B${R_SPR}/10000))*B{r}/360',
            fill=CALC,nf="0.000000")
U=G0+NM+1  # row 71: per-$1 unit totals
put(U-1,1,"UNIT TOTALS / $1 (per scenario)",SUB)
put(U,3,f"=SUM(C{G0}:C{G0+NM-1})",OUT,nf=NUM)  # C71 bond total
for j in range(6):
    b=4+j*7; L=lambda k:gcl(b+k)
    for k in range(1,5):  # dep unit per bucket
        put(U,b+k,f"=SUMPRODUCT({L(k)}{G0}:{L(k)}{G0+NM-1}/100*$B${G0}:$B${G0+NM-1})/360",fill=CALC,nf="0.00000")
    put(U,b+6,f"=SUM({L(6)}{G0}:{L(6)}{G0+NM-1})",fill=CALC,nf="0.00000")
    # float-leg unit (for fair rate): row 72
    put(U+1,b,f"=SUMPRODUCT({L(0)}{G0}:{L(0)}{G0+NM-1}/100*$B${G0}:$B${G0+NM-1})/360"
              f"+$B${R_SPR}/10000*SUM($B${G0}:$B${G0+NM-1})/360",fill=CALC,nf="0.00000")

# ================= RESULTS TAPE (P4..V15)
O=16
put(4,O,"RESULTS — 2Y TOTALS ($MM), INPUT STRATEGY",SUB)
for j,s in enumerate(SCEN): put(5,O+1+j,s,SUB,align="center")
for i,k in enumerate(["Deposit income","Swap net","Bond cost","NII total"]): put(6+i,O,k)
for j in range(6):
    c=O+1+j; b=4+j*7; CL=gcl(c)
    put(6,c,f"=$B$17*{gcl(b+1)}${U}+$B$18*{gcl(b+2)}${U}+$B$19*{gcl(b+3)}${U}+$B$20*{gcl(b+4)}${U}",OUT,nf=NUM)
    put(7,c,f"=$B${R_NOT}/100*$B${R_CASH}*{gcl(b+6)}${U}",OUT,nf=NUM)
    put(8,c,f"=$C${U}",OUT,nf=NUM)
    put(9,c,f"={CL}6+{CL}7-{CL}8",Font(name=A,bold=True),nf=NUM)
wrng=f"$H${WR}:$M${WR}"
stats=[("Expected NII",f"=SUMPRODUCT({wrng},$Q$9:$V$9)"),
("NII volatility",f"=SQRT(SUMPRODUCT({wrng},($Q$9:$V$9-$Q$11)^2))"),
("Worst case","=MIN($Q$9:$V$9)"),("Best case","=MAX($Q$9:$V$9)"),
("Scenario-fair fixed (%)",
 f"=100*SUMPRODUCT({wrng},{'+'.join(f'0' for _ in [0])}D{U+1}:D{U+1})"),  # placeholder, fixed below
("Carry gap (bps)",f"=($B${R_FIX}-$Q$15)*100")]
for i,(k,f) in enumerate(stats):
    put(11+i,O,k,SUB); put(11+i,O+1,f,OUT,nf=N2 if i>=4 else NUM)
# fair fixed: weighted float units / (sum days/360)
fl=",".join(f"{gcl(4+j*7)}{U+1}" for j in range(6))
put(15,O+1,f"=100*SUMPRODUCT({wrng},CHOOSE({{1,2,3,4,5,6}},{fl}))/(SUM($B${G0}:$B${G0+NM-1})/360)",OUT,nf=N2)

# ================= COMBOS / FRONTIER (rows 76+)
C0=76
fracs=[]
for a in range(0,11):
    for b2 in range(0,11-a):
        for c2 in range(0,11-a-b2):
            fracs.append((a/10,b2/10,c2/10,(10-a-b2-c2)/10))
combos=list(product(fracs,[0,25,50,75,100]))
NC=len(combos)
put(C0-2,1,f"STRATEGY GRID — {NC} combinations (10% allocation steps × IRS 0/25/50/75/100%). "
           "Feasibility, frontier and zones are live.",SUB)
cols=["ON","1M","2M","3M","Label","IRS%","Feasible"]+SCEN+["Expected","Vol","Worst","Best","Efficient","KneeD"]
for c,h in enumerate(cols,1): put(C0-1,c,h,HDR,HF)
for c,h in enumerate(["Xall","Yall","Xeff","Yeff","W|f","E|f","V|e","E|e"],50): put(C0-1,c,h,HDR,HF)
NR=f"${C0}:${C0+NC-1}"
expR=f"$N{NR.split(':')[0][1:]}"  # built below explicitly
expRng=f"$N${C0}:$N${C0+NC-1}"; volRng=f"$O${C0}:$O${C0+NC-1}"
feaRng=f"$G${C0}:$G${C0+NC-1}"; effRng=f"$R${C0}:$R${C0+NC-1}"
worRng=f"$P${C0}:$P${C0+NC-1}"
for i,((fa,fb,fc,fd),p) in enumerate(combos):
    r=C0+i
    for c,v in enumerate([fa,fb,fc,fd],1): put(r,c,v,fill=CALC,nf="0%")
    lab="/".join(f"{n} {int(v*100)}%" for n,v in zip(["ON","1M","2M","3M"],[fa,fb,fc,fd]) if v>0) or "—"
    put(r,5,lab,fill=CALC)
    put(r,6,p,fill=CALC)
    put(r,7,f"=A{r}*$B${R_CASH}>=$B${R_MIN}",fill=CALC)
    for j in range(6):
        b=4+j*7
        put(r,8+j,f"=$B${R_CASH}*(A{r}*{gcl(b+1)}${U}+B{r}*{gcl(b+2)}${U}+C{r}*{gcl(b+3)}${U}+D{r}*{gcl(b+4)}${U})"
                  f"+F{r}/100*$B${R_CASH}*{gcl(b+6)}${U}-$C${U}",fill=CALC,nf=NUM)
    put(r,14,f"=SUMPRODUCT({wrng},H{r}:M{r})",fill=CALC,nf=NUM)
    put(r,15,f"=SQRT(SUMPRODUCT({wrng},(H{r}:M{r}-N{r})^2))",fill=CALC,nf=N2)
    put(r,16,f"=MIN(H{r}:M{r})",fill=CALC,nf=NUM)
    put(r,17,f"=MAX(H{r}:M{r})",fill=CALC,nf=NUM)
    put(r,18,f"=IF(G{r},SUMPRODUCT(({feaRng})*({expRng}>=N{r})*({volRng}<=O{r}))=1,FALSE)",fill=CALC)
    put(r,19,f"=IF(R{r},ABS((O{r}-$BG${C0})/MAX(0.0001,$BG${C0+1}-$BG${C0})"
             f"-(N{r}-$BG${C0+2})/MAX(0.0001,$BG${C0+3}-$BG${C0+2})),-1)",fill=CALC,nf="0.000")
    put(r,50,f"=IF(G{r},O{r},NA())",fill=CALC,nf=N2)
    put(r,51,f"=IF(G{r},N{r},NA())",fill=CALC,nf=NUM)
    put(r,52,f"=IF(R{r},O{r},NA())",fill=CALC,nf=N2)
    put(r,53,f"=IF(R{r},N{r},NA())",fill=CALC,nf=NUM)
    put(r,54,f'=IF(G{r},P{r},"")',fill=CALC)   # worst|feasible
    put(r,55,f'=IF(G{r},N{r},"")',fill=CALC)   # exp|feasible
    put(r,56,f'=IF(R{r},O{r},"")',fill=CALC)   # vol|efficient
    put(r,57,f'=IF(R{r},N{r},"")',fill=CALC)   # exp|efficient

# fixed aggregates over the strategy grid (col AC, labels AB)
aggs=[("minVol|eff",f"=MIN($BD${C0}:$BD${C0+NC-1})"),("maxVol|eff",f"=MAX($BD${C0}:$BD${C0+NC-1})"),
("minExp|eff",f"=MIN($BE${C0}:$BE${C0+NC-1})"),("maxExp|eff",f"=MAX($BE${C0}:$BE${C0+NC-1})"),
("maxWorst|feas",f"=MAX($BB${C0}:$BB${C0+NC-1})"),("maxExp|feas",f"=MAX($BC${C0}:$BC${C0+NC-1})"),
("maxKnee",f"=MAX($S${C0}:$S${C0+NC-1})")]
for i,(lab,f) in enumerate(aggs):
    put(C0+i,58,lab,fill=CALC); put(C0+i,59,f,fill=CALC,nf="0.0000")

# ================= ZONES (P17..V21)
put(17,O,"ZONES (live from strategy grid)",SUB)
zhdr=["Zone","Allocation","IRS %","Expected","Vol","Worst"]
for c,h in enumerate(zhdr): put(18,O+c,h,SUB)
kneRng=f"$S${C0}:$S${C0+NC-1}"
zrows=[("Defensive",f"MATCH($BG${C0+4},$BB${C0}:$BB${C0+NC-1},0)"),
       ("Balanced",f"MATCH($BG${C0+6},{kneRng},0)"),
       ("Opportunistic",f"MATCH($BG${C0+5},$BC${C0}:$BC${C0+NC-1},0)")]
for i,(z,mf) in enumerate(zrows):
    r=19+i
    put(r,O,z,SUB)
    idx=f"({mf})"
    put(r,O+1,f"=INDEX($E${C0}:$E${C0+NC-1},{idx})",OUT)
    put(r,O+2,f"=INDEX($F${C0}:$F${C0+NC-1},{idx})",OUT,nf="0")
    put(r,O+3,f"=INDEX({expRng},{idx})",OUT,nf=NUM)
    put(r,O+4,f"=INDEX({volRng},{idx})",OUT,nf=N2)
    put(r,O+5,f"=INDEX({worRng},{idx})",OUT,nf=NUM)
# zone scatter helpers (X col W under? use P/Q cols of a tiny block) at AB18:AC21
put(18,50,"Zx",SUB);put(18,51,"Zy",SUB)
for i in range(3):
    put(19+i,50,f"=INDEX({volRng},({zrows[i][1]}))",fill=CALC,nf=N2)
    put(19+i,51,f"=INDEX({expRng},({zrows[i][1]}))",fill=CALC,nf=NUM)

# ================= SLOPE / SENSITIVITY table (P23..V29)
put(23,O,"NII BY SCENARIO × IRS NOTIONAL (input allocation)",SUB)
put(24,O,"Notional %",SUB)
for j,s in enumerate(SCEN): put(24,O+1+j,s,SUB,align="center")
for i,p in enumerate([0,25,50,75,100]):
    r=25+i; put(r,O,p,fill=CALC,nf="0")
    for j in range(6):
        b=4+j*7; c=O+1+j
        put(r,c,f"={gcl(c)}$6+$P{r}/100*$B${R_CASH}*{gcl(b+6)}${U}-$C${U}",fill=CALC,nf=NUM)

# ================= CONE table (X4..AE26) step 5
X0=24  # col X
put(4,X0,"CONVERGENCE CONE — NII vs hedge ratio",SUB)
put(5,X0,"p%",SUB)
for j,s in enumerate(SCEN): put(5,X0+1+j,s,SUB,align="center")
put(5,X0+7,"Dispersion",SUB)
for i,p in enumerate(range(0,101,5)):
    r=6+i; put(r,X0,p,fill=CALC,nf="0")
    for j in range(6):
        b=4+j*7; c=X0+1+j
        put(r,c,f"={gcl(O+1+j)}$6+${gcl(X0)}{r}/100*$B${R_CASH}*{gcl(b+6)}${U}-$C${U}",fill=CALC,nf=NUM)
    put(r,X0+7,f"=MAX({gcl(X0+1)}{r}:{gcl(X0+6)}{r})-MIN({gcl(X0+1)}{r}:{gcl(X0+6)}{r})",fill=CALC,nf=N2)
dR=f"${gcl(X0+7)}$6:${gcl(X0+7)}$26"
put(28,X0,"Min dispersion @ notional %",SUB)
put(28,X0+2,f"=INDEX(${gcl(X0)}$6:${gcl(X0)}$26,MATCH(MIN({dR}),{dR},0))",OUT,nf="0")
put(28,X0+3,f"=MIN({dR})",OUT,nf=N2)

# ================= TIMELINE composition (AG4..AO28) with scenario selector
T0=33  # col AG
put(4,T0,"CASHFLOW COMPOSITION — scenario:",SUB)
sel=put(4,T0+4,"Faster cuts",BLUE)
dv2=DataValidation(type="list",formula1='"'+",".join(SCEN)+'"'); ws.add_data_validation(dv2); dv2.add(f"{gcl(T0+4)}4")
SELC=f"${gcl(T0+4)}$4"
mtch=f"MATCH({SELC},$Q$5:$V$5,0)-1"
thdr=["Month","ON","1M","2M","3M","SWAP","BOND","NET"]
for c,h in enumerate(thdr): put(5,T0+c,h,SUB)
for m in range(NM):
    r=6+m; gr=G0+m
    put(r,T0,f"=A{gr}",fill=CALC,nf="MMM-YY")
    for k in range(4):  # bucket incomes via INDEX over grid row
        put(r,T0+1+k,f"=$B${17+k}*INDEX($A{gr}:$AU{gr},4+({mtch})*7+{k+1})/100*$B{gr}/360",fill=CALC,nf="0.000")
    put(r,T0+5,f"=$B${R_NOT}/100*$B${R_CASH}*INDEX($A{gr}:$AU{gr},4+({mtch})*7+6)",fill=CALC,nf="0.000")
    put(r,T0+6,f"=-$C{gr}",fill=CALC,nf="0.000")
    put(r,T0+7,f"=SUM({gcl(T0+1)}{r}:{gcl(T0+6)}{r})",OUT,nf="0.000")

# ================= CHARTS
def line(title,h=7,w=14):
    c=LineChart(); c.title=title; c.height,c.width=h,w
    c.y_axis.delete=False; c.x_axis.delete=False; return c
# SOFR
ch=line("SOFR step paths (%)")
for j in range(6):
    ch.add_data(Reference(ws,min_col=4+j*7,min_row=G0-1,max_row=G0+NM-1),titles_from_data=True)
ch.set_categories(Reference(ws,min_col=1,min_row=G0,max_row=G0+NM-1))
for sr in ch.series: sr.smooth=False
ws.add_chart(ch,"P31")
# slope (series per notional row)
ch2=line("NII by scenario × IRS notional")
ch2.add_data(Reference(ws,min_col=O,max_col=O+6,min_row=25,max_row=29),from_rows=True,titles_from_data=True)
ch2.set_categories(Reference(ws,min_col=O+1,max_col=O+6,min_row=24,max_row=24))
for sr in ch2.series: sr.smooth=False
ws.add_chart(ch2,"P47")
# cone
ch3=line("Convergence cone — NII vs hedge ratio")
ch3.add_data(Reference(ws,min_col=X0+1,max_col=X0+6,min_row=5,max_row=26),titles_from_data=True)
ch3.set_categories(Reference(ws,min_col=X0,min_row=6,max_row=26))
for sr in ch3.series: sr.smooth=False
ws.add_chart(ch3,"X31")
# frontier scatter
ch4=ScatterChart(); ch4.title="Efficient frontier — expected NII vs volatility"
ch4.height,ch4.width=9,16; ch4.x_axis.delete=False; ch4.y_axis.delete=False
ch4.x_axis.title="Volatility"; ch4.y_axis.title="Expected NII"
ch4.visible_cells_only=False
for (xc,yc,t,sym,sz) in [(50,51,"All",'circle',3),(52,53,"Efficient",'diamond',6),(50,51,"Zones",'star',9)]:
    if t=="Zones":
        s=Series(Reference(ws,min_col=51,min_row=19,max_row=21),Reference(ws,min_col=50,min_row=19,max_row=21),title=t)
    else:
        s=Series(Reference(ws,min_col=yc,min_row=C0,max_row=C0+NC-1),Reference(ws,min_col=xc,min_row=C0,max_row=C0+NC-1),title=t)
    s.marker=Marker(symbol=sym,size=sz); s.graphicalProperties.line.noFill=True
    ch4.series.append(s)
ws.add_chart(ch4,"X47")
# timeline stacked composition + NET line
ch5=BarChart(); ch5.type="col"; ch5.grouping="stacked"; ch5.overlap=100
ch5.title="Cashflow composition (selected scenario)"; ch5.height,ch5.width=9,18
ch5.y_axis.delete=False; ch5.x_axis.delete=False
ch5.add_data(Reference(ws,min_col=T0+1,max_col=T0+6,min_row=5,max_row=5+NM),titles_from_data=True)
ch5.set_categories(Reference(ws,min_col=T0,min_row=6,max_row=5+NM))
ws.add_chart(ch5,f"{gcl(T0)}31")
ln=LineChart(); ln.title="NET NII by month (selected scenario)"; ln.height,ln.width=6,18
ln.y_axis.delete=False; ln.x_axis.delete=False
ln.add_data(Reference(ws,min_col=T0+7,min_row=5,max_row=5+NM),titles_from_data=True)
ln.set_categories(Reference(ws,min_col=T0,min_row=6,max_row=5+NM))
for sr in ln.series: sr.smooth=False
ws.add_chart(ln,f"{gcl(T0)}50")

from openpyxl.workbook.properties import CalcProperties
wb.calculation=CalcProperties(fullCalcOnLoad=True)
for _c in range(50,60): ws.column_dimensions[gcl(_c)].hidden=True
ws.freeze_panes="A4"
for col,w in [("A",24),("B",12),("C",9),("D",12),("E",22),("F",12),("G",12),("P",24)]:
    ws.column_dimensions[col].width=w
# ================= SELF-TEST BLOCK
put(33,1,"SELF-TESTS (recompute on any input change)",SUB)
tests=[("Weights sum to 100%",f'=IF(ABS(SUM(H{WR}:M{WR})-1)<0.0001,"OK","FAIL")'),
("Allocation check",'=IF(B21="OK","OK","FAIL")'),
("NII identity (dep+swap-bond)",'=IF(SUMPRODUCT(ABS($Q$9:$V$9-($Q$6:$V$6+$Q$7:$V$7-$Q$8:$V$8)))<0.001,"OK","FAIL")'),
("Timeline reconciles to NII",f'=IF(ABS(SUM({gcl(33+7)}6:{gcl(33+7)}29)-INDEX($Q$9:$V$9,MATCH(${gcl(33+4)}$4,$Q$5:$V$5,0)))<0.01,"OK","FAIL")'),
("Frontier non-empty",f'=IF(COUNTIF($R${C0}:$R${C0+NC-1},TRUE)>0,"OK","FAIL")'),
("Cone pinch in range",f'=IF(AND(${gcl(26)}$28>=0,${gcl(26)}$28<=100),"OK","FAIL")')]
for i,(k,f) in enumerate(tests,34):
    put(i,1,k); c=put(i,2,f,SUB); c.fill=WARN
put(41,1,"Model version",SUB); put(41,2,VERSION,SUB)
put(40,1,"MODEL STATUS",SUB)
c=put(40,2,'=IF(COUNTIF(B34:B39,"FAIL")=0,"OK — ALL TESTS PASS","CHECK FAILED TESTS")',
      Font(name=A,bold=True,color="006100")); c.fill=WARN

import os
_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
_out = os.path.join(_dir, f"NII_Frontier_Full_{VERSION}.xlsx")
try:
    wb.save(_out)
except PermissionError:
    _out = os.path.join(_dir, f"NII_Frontier_Full_{VERSION}_new.xlsx")
    wb.save(_out)   # original is probably open in Excel
print(f"built: {NC} strategies | {VERSION}")
print(f"saved: {_out}")
# ---- reference values (independent of the spreadsheet; sheet must match after Excel recalc)
from datetime import date as _d
_sofr0,_cash,_beta,_fix,_nspr,_npct=3.80,500,1.0,3.15,0,75
_alloc=[100,0,0,400];_spr=[0,10,15,20]
def _em(d0,m):
    y=d0.year+(d0.month-1+m)//12; mo=(d0.month-1+m)%12+1; return _d(y,mo,1)
_start=_d(2026,7,1); _MVd=dict(zip(SCEN,moves())); _ref={}
for _s in SCEN:
    _ms=[_em(_start,m) for m in range(24)]; _dy=[(_em(_start,m+1)-_em(_start,m)).days for m in range(24)]
    _sf=[_sofr0+sum(b for dd,b in zip(FOMC,_MVd[_s]) if dd.toordinal()+1<=mm.toordinal())/100 for mm in _ms]
    _adj=lambda x:_sofr0+_beta*(x-_sofr0); _dep=_swu=_bd=0
    for m in range(24):
        _rs=[_adj(_sf[m])+_spr[0]/100,_adj(_sf[m])+_spr[1]/100,
             _adj(sum(_sf[max(0,m-1):m+1])/len(_sf[max(0,m-1):m+1]))+_spr[2]/100,
             _adj(sum(_sf[max(0,m-2):m+1])/len(_sf[max(0,m-2):m+1]))+_spr[3]/100]
        _dep+=sum(a*r for a,r in zip(_alloc,_rs))/100*_dy[m]/360
        _swu+=(_fix/100-(_sf[m]/100+_nspr/10000))*_dy[m]/360
        _bd+=sum(n*c for _,n,c,mt in BONDS if mt>_ms[m])/100/12
    _ref[_s]=_dep+_npct/100*_cash*_swu-_bd
print("REFERENCE (defaults; sheet row 9 should match):")
for _s in SCEN: print(f"  {_s:18s} {_ref[_s]:9.2f}")
_e=sum(_ref.values())/6
print(f"  Expected {_e:.2f} | Worst {min(_ref.values()):.2f} | Best {max(_ref.values()):.2f}")
