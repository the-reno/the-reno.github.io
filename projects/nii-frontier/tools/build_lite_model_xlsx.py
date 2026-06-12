"""NII FRONTIER LITE v4.0 - generic-inputs standalone spreadsheet.
Single tab, live formulas, no macros (VBA ships separately as nii_tools.txt).

v4.0:
- Analysis tenor is an input (6-36 months); the grid truncates automatically (days=0 past horizon)
- 8 scenario slots (names, dates, moves, scale, weight all editable; weight 0 disables a slot)
- Up to 30 FOMC dates and 12 bonds; blank rows ignored
- Deposit interest ACCRUES: the cash balance compounds month over month, per scenario
- Efficiency chart: volatility vs expected NII along the hedge ratio 0-100%
- Expense chart: stacked area, one band per bond issue, composing total interest cost
Run: python3 build_lite_model_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.chart import LineChart, BarChart, AreaChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.properties import CalcProperties
from openpyxl.utils import get_column_letter as gcl
from datetime import date

VERSION="v4.0"
A="Arial"
NAVY="1F3864"; AMBER="BF8F00"; GREY="F2F2F2"
BLUE=Font(name=A,color="0000FF"); BLK=Font(name=A,size=10); SUB=Font(name=A,bold=True,size=10)
HDR=Font(name=A,bold=True,color="FFFFFF",size=10); HF=PatternFill("solid",start_color=NAVY)
CALC=PatternFill("solid",start_color=GREY); OUT=Font(name=A,bold=True,color=NAVY,size=10)
WARN=PatternFill("solid",start_color="FFF2CC")
THIN=Side(style="thin",color="BFBFBF")
BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
NUM='#,##0.0;(#,##0.0);""'; N2='#,##0.00;(#,##0.00);""'; N0='0;-0;"-"'

# ---------------- data
NS,NB,NF,NMX=8,12,30,36
SCEN=["Flat","Gradual cuts","Faster cuts","Deep cuts","Delayed cuts","Higher for longer","Hike","Aggressive hikes"]
FOMC=[date(2026,6,17),date(2026,7,29),date(2026,9,16),date(2026,10,28),date(2026,12,9),
date(2027,1,27),date(2027,3,17),date(2027,4,28),date(2027,6,16),date(2027,7,28),
date(2027,9,15),date(2027,10,27),date(2027,12,8),date(2028,1,26),date(2028,3,15),
date(2028,4,26),date(2028,6,14),date(2028,7,26),date(2028,9,13),date(2028,11,1),date(2028,12,13)]
def MVS():
    z=lambda mv:[mv.get(i,0) for i in range(len(FOMC))]
    return [z({}),z({i:-25 for i in range(0,12,2)}),z({i:-25 for i in range(6)}),
            z({i:-50 for i in range(6)}),z({i:-25 for i in range(4,10)}),
            z({8:-25,9:-25}),z({0:25,1:25}),z({i:25 for i in range(4)})]
BONDS=[("3.50% '27",476.4,3.50,date(2027,9,15)),("6.60% '28",109.9,6.60,date(2028,7,15)),
("3.90% '29",900.0,3.90,date(2029,11,19)),("4.65% '31",400.0,4.65,date(2031,3,12)),
("6.05% '34",500.0,6.05,date(2034,5,14)),("6.35% '40",500.0,6.35,date(2040,3,15)),
("5.10% '44",300.0,5.10,date(2044,5,15))]

# ---------------- layout map
R_START,R_TEN,R_SOFR0,R_CASH,R_BETA,R_MIN,R_DIR,R_FIX,R_SPR,R_NOT=range(5,15)
AL0=18          # allocation rows 18-21, check 22
BH,B0=25,26     # bonds header 25, data 26..37
B1=B0+NB-1
T0r,T1r=40,46   # self-tests 41-45, status 46, version 47
FR0,FR1=6,6+NF-1        # FOMC rows 6..35 (cols F,G + H..O)
SCr,WRr=FR1+2,FR1+3      # scale 37, weights 38
G0=51; GH=G0-1; GL=G0-2  # grid label 49, header 50, data 51..86
U=G0+NMX+1               # unit totals row 88 (87 blank)
# grid columns: A month, B days, C total bond, D..O per-bond (12), scenario blocks from col 16
SB=16; BW=8              # block: SOFR,ON,1M,2M,3M,Bal,Dep,Swap/$
RQ=17                    # results label col Q, values R..Y (18..25)
SEN0=19                  # sensitivity rows 19..29 (p 0..100), header 18, eff cols
TLc=31                   # timeline cols AE.. (31..39), rows 6..41
CH0=92                   # charts dashboard from row 92

wb=Workbook(); ws=wb.active; ws.title="MODEL"
ws.sheet_view.showGridLines=False
def put(r,c,v,font=BLK,fill=None,nf=None,align=None,note=None,border=False):
    x=ws.cell(row=r,column=c,value=v); x.font=font
    if fill:x.fill=fill
    if nf:x.number_format=nf
    if align:x.alignment=Alignment(horizontal=align)
    if note:x.comment=Comment("IMPACT: "+note,"RONIN",height=120,width=270)
    if border:x.border=BOX
    return x
def banner(r,c0,c1,text):
    ws.merge_cells(start_row=r,start_column=c0,end_row=r,end_column=c1)
    x=ws.cell(row=r,column=c0,value=text); x.font=HDR; x.fill=HF
    x.alignment=Alignment(horizontal="left",indent=1)

put(1,1,f"NII EFFICIENT FRONTIER — LITE {VERSION}",Font(name=A,bold=True,size=15,color=NAVY))
put(2,1,"Single tab · live formulas · no macros. Blue cells are inputs — hover any for its impact. "
        "Blank FOMC/bond rows are ignored; set a scenario's weight to 0 to disable it. "
        "Deposit interest accrues into the balance (compounding). Monthly grid, ACT/360, bonds 30/360. "
        "Illustrative only.",Font(name=A,size=9,color="595959"))

# ================= INPUTS
banner(4,1,4,"INPUTS")
ins=[("Start month (1st)",date(2026,7,1),"YYYY-MM-DD",
  "Anchors the monthly grid; every date-driven calculation shifts with it."),
("Analysis tenor (months, 6-36)",24,N0,
  "Length of the horizon. The grid truncates automatically - income, swap and bond cost stop accruing past it."),
("Initial SOFR (%)",3.80,N2,
  "Day-0 rate; sets the level of every scenario path and floating leg."),
("Total cash ($mm)",500.0,NUM,
  "Opening balance. Interest accrues into it, so later months earn on a larger base. IRS notional is % of this initial amount."),
("Deposit beta",1.0,N2,
  "Pass-through of SOFR moves to deposit rates. Below 1.0 dampens scenario differences on the cash side."),
("Min ON balance ($mm)",100.0,NUM,
  "Liquidity floor held overnight. Raising it shortens the book: faster repricing, more cut exposure."),
("Swap direction (RCV/PAY)","RCV",None,
  "RCV fixed gains when rates fall (hedges this asset-sensitive book); PAY is the mirror."),
("Swap fixed rate (%)",3.15,N2,
  "Traded mid. Its gap to the scenario-fair rate (output) is the carry cost of hedging - it shapes the whole analysis."),
("Swap float spread (bps)",0.0,NUM,
  "Added to SOFR on the floating leg; positive spread reduces receive-fixed value bp-for-bp."),
("IRS notional (% of cash)",75.0,NUM,
  "The hedge ratio. See the efficiency and cone charts for what each level buys."),
]
for i,(k,v,nf,nt) in enumerate(ins,5):
    put(i,1,k); put(i,2,v,BLUE,nf=nf,note=nt,border=True)
dv=DataValidation(type="list",formula1='"RCV,PAY"'); ws.add_data_validation(dv); dv.add(f"B{R_DIR}")

banner(16,1,4,"ALLOCATION ($mm) / SPREAD (bps)")
put(17,1,"Bucket",SUB);put(17,2,"$mm",SUB);put(17,3,"Spread",SUB)
anotes=["Overnight: reprices daily - first hurt by cuts, first helped by hikes.",
"1M rolling: ~1 month repricing lag.","2M rolling: ~2 month lag.",
"3M rolling: slowest repricing - most income protection in cuts."]
for i,((b,al,sp),an) in enumerate(zip([("ON",100,0),("1M",0,10),("2M",0,15),("3M",400,20)],anotes)):
    r=AL0+i
    put(r,1,b); put(r,2,al,BLUE,nf=NUM,note=an,border=True)
    put(r,3,sp,BLUE,nf=NUM,border=True,
        note="Margin over SOFR in this bucket; raising it favors this tenor in every scenario.")
c=put(22,1,"Check"); c=put(22,2,f'=IF(SUM(B{AL0}:B{AL0+3})<>B{R_CASH},"SUM<>CASH",IF(B{AL0}<B{R_MIN},"ON<MIN","OK"))',SUB); c.fill=WARN

banner(24,1,4,"FIXED-RATE DEBT (blank rows ignored)")
put(BH,1,"Issue",SUB);put(BH,2,"$mm",SUB);put(BH,3,"Cpn %",SUB);put(BH,4,"Maturity",SUB)
bn="Each issue is a band in the expense chart. Constant across scenarios: sets the NII level, not its dispersion."
for i in range(NB):
    r=B0+i
    if i<len(BONDS):
        n,nm,cp,mt=BONDS[i]
        put(r,1,n,border=True); put(r,2,nm,BLUE,nf=NUM,note=bn if i==0 else None,border=True)
        put(r,3,cp,BLUE,nf=N2,border=True); put(r,4,mt,BLUE,nf="YYYY-MM-DD",border=True)
    else:
        for cc,nf in [(1,None),(2,NUM),(3,N2),(4,"YYYY-MM-DD")]:
            put(r,cc,None,BLUE,nf=nf,border=True)

# ================= SELF-TESTS
banner(T0r,1,4,"SELF-TESTS")
# (formulas appended after dependent blocks exist - see end)

# ================= FOMC x SCENARIOS
banner(4,6,6+1+NS,"FED SCENARIOS — bps per decision · names, dates, moves, scale and weights all editable")
put(5,6,"Decision",SUB); put(5,7,"Effective",SUB)
for j,s in enumerate(SCEN):
    put(5,8+j,s,BLUE,align="center",border=True,
        note="Scenario name; flows to every header and chart legend. Set its weight to 0 to disable.")
MV=MVS()
for i in range(NF):
    r=FR0+i
    put(r,6,FOMC[i] if i<len(FOMC) else None,BLUE,nf="YYYY-MM-DD",border=True,
        note="FOMC decision date; the move is effective next day. Blank = ignored." if i==0 else None)
    put(r,7,f'=IF(F{r}="","",F{r}+1)',fill=CALC,nf="YYYY-MM-DD")
    for j in range(NS):
        put(r,8+j,(MV[j][i] if i<len(FOMC) else None),BLUE,nf=N0,align="center",border=True,
            note="Move in bps at this meeting under this scenario." if (i==0 and j==0) else None)
put(SCr,6,"Scale ×",SUB)
put(WRr,6,"Weight",SUB)
for j in range(NS):
    put(SCr,8+j,1.0,BLUE,nf="0.00",align="center",border=True,
        note="Multiplies every move in this column - 2.0 doubles the path, 0.5 halves it. Cheap scenario expansion.")
    put(WRr,8+j,1/len(SCEN),BLUE,nf="0.0%",align="center",border=True,
        note="Probability. Drives expected NII and volatility; 0 removes the scenario from worst/best too. Sum to 100%.")

# ================= CALC GRID
banner(GL,1,SB+NS*BW-1,"CALCULATION GRID — monthly, per scenario (collapsible · do not edit)")
ghdr=["Month","Days","Bond Σ"]+[f"=IF($A${B0+i}=0,\"\",$A${B0+i})" for i in range(NB)]
for c,h in enumerate(ghdr,1):
    x=ws.cell(row=GH,column=c,value=h); x.font=HDR; x.fill=HF
for j in range(NS):
    b=SB+j*BW
    put(GH,b,f"={gcl(8+j)}5",HDR,HF)
    for k,h in enumerate(["ON%","1M%","2M%","3M%","Balance","Dep","Swap/$"],1):
        put(GH,b+k,h,HDR,HF)
for m in range(NMX):
    r=G0+m
    ws.row_dimensions[r].outlineLevel=1
    put(r,1,f"=IF({m}<$B${R_TEN},EDATE($B${R_START},{m}),\"\")",fill=CALC,nf="MMM-YY")
    put(r,2,f"=IF(A{r}=\"\",0,EDATE(A{r},1)-A{r})",fill=CALC)
    put(r,3,f"=SUM(D{r}:{gcl(3+NB)}{r})",fill=CALC,nf=N2)
    for i in range(NB):
        cc=4+i; br=B0+i
        put(r,cc,f"=IF($B{r}=0,0,IF(AND($B${br}>0,$D${br}>$A{r}),$B${br}*$C${br}/100/12,0))",fill=CALC,nf="0.000")
    for j in range(NS):
        b=SB+j*BW; L=lambda k:gcl(b+k); mc=gcl(8+j)
        sofr=(f"$B${R_SOFR0}+{mc}${SCr}*SUMIFS({mc}${FR0}:{mc}${FR1},$G${FR0}:$G${FR1},\"<=\"&"
              f"IF($A{r}=\"\",EDATE($B${R_START},{m}),$A{r}),$F${FR0}:$F${FR1},\"<>\")/100")
        put(r,b,f"={sofr}",fill=CALC,nf=N2)
        adj=lambda s:f"$B${R_SOFR0}+$B${R_BETA}*({s}-$B${R_SOFR0})"
        put(r,b+1,f"={adj(L(0)+str(r))}+$C${AL0}/100",fill=CALC,nf=N2)
        put(r,b+2,f"={adj(L(0)+str(r))}+$C${AL0+1}/100",fill=CALC,nf=N2)
        col=f"{L(0)}${G0}:{L(0)}${G0+NMX-1}"
        put(r,b+3,f"={adj(f'AVERAGE(INDEX({col},MAX(1,{m})):INDEX({col},{m+1}))')}+$C${AL0+2}/100",fill=CALC,nf=N2)
        put(r,b+4,f"={adj(f'AVERAGE(INDEX({col},MAX(1,{m-1})):INDEX({col},{m+1}))')}+$C${AL0+3}/100",fill=CALC,nf=N2)
        if m==0:
            put(r,b+5,f"=$B${R_CASH}",fill=CALC,nf=NUM)
        else:
            put(r,b+5,f"={L(5)}{r-1}+{L(6)}{r-1}",fill=CALC,nf=NUM)
        blend=(f"($B${AL0}*{L(1)}{r}+$B${AL0+1}*{L(2)}{r}+$B${AL0+2}*{L(3)}{r}+$B${AL0+3}*{L(4)}{r})/$B${R_CASH}")
        put(r,b+6,f"={L(5)}{r}*({blend})/100*$B{r}/360",fill=CALC,nf="0.0000")
        put(r,b+7,f'=IF($B${R_DIR}="RCV",1,-1)*(($B${R_FIX}/100)-({L(0)}{r}/100+$B${R_SPR}/10000))*$B{r}/360',
            fill=CALC,nf="0.000000")
put(U-1,1,"UNIT TOTALS",SUB)
for j in range(NS):
    b=SB+j*BW; L=lambda k:gcl(b+k)
    put(U,b+6,f"=SUM({L(6)}{G0}:{L(6)}{G0+NMX-1})",fill=CALC,nf=NUM)           # dep total
    put(U,b+7,f"=SUM({L(7)}{G0}:{L(7)}{G0+NMX-1})",fill=CALC,nf="0.00000")     # swap unit
    put(U+1,b,f"=SUMPRODUCT({L(0)}{G0}:{L(0)}{G0+NMX-1}/100*$B${G0}:$B${G0+NMX-1})/360"
              f"+$B${R_SPR}/10000*SUM($B${G0}:$B${G0+NMX-1})/360",fill=CALC,nf="0.00000")  # float unit
put(U,3,f"=SUM(C{G0}:C{G0+NMX-1})",fill=CALC,nf=NUM)

# ================= RESULTS
banner(4,RQ,RQ+NS+1,"RESULTS — HORIZON TOTALS ($MM)")
for j in range(NS): put(5,RQ+1+j,f"={gcl(8+j)}5",SUB,align="center")
for i,k in enumerate(["Deposit income","Swap net","Bond cost","NII total"]): put(6+i,RQ,k)
wrng=f"$H${WRr}:${gcl(7+NS)}${WRr}"
wsum=f"SUM({wrng})"
for j in range(NS):
    c=RQ+1+j; b=SB+j*BW; CL=gcl(c)
    put(6,c,f"={gcl(b+6)}${U}",OUT,nf=NUM)
    put(7,c,f"=$B${R_NOT}/100*$B${R_CASH}*{gcl(b+7)}${U}",OUT,nf=NUM)
    put(8,c,f"=$C${U}",OUT,nf=NUM)
    put(9,c,f"={CL}6+{CL}7-{CL}8",Font(name=A,bold=True),nf=NUM)
    put(10,c,f'=IF({gcl(8+j)}${WRr}=0,"",{CL}9)',fill=CALC,nf=NUM)   # active helper (weight>0)
nii=f"$R$9:${gcl(RQ+NS)}$9"; act=f"$R$10:${gcl(RQ+NS)}$10"
put(12,RQ,"Expected NII",SUB); put(12,RQ+1,f"=SUMPRODUCT({wrng},{nii})/MAX(0.0001,{wsum})",OUT,nf=NUM)
put(13,RQ,"NII volatility",SUB)
put(13,RQ+1,f"=SQRT(MAX(0,SUMPRODUCT({wrng},({nii}-$R$12)^2)/MAX(0.0001,{wsum})))",OUT,nf=N2)
put(14,RQ,"Worst / Best",SUB); put(14,RQ+1,f"=MIN({act})",OUT,nf=NUM); put(14,RQ+2,f"=MAX({act})",OUT,nf=NUM)
fl=",".join(f"{gcl(SB+j*BW)}{U+1}" for j in range(NS))
put(15,RQ,"Scenario-fair fixed (%)",SUB)
put(15,RQ+1,f"=100*SUMPRODUCT({wrng},CHOOSE({{1,2,3,4,5,6,7,8}},{fl}))/MAX(0.0001,{wsum})"
            f"/(SUM($B${G0}:$B${G0+NMX-1})/360)",OUT,nf=N2)
put(16,RQ,"Carry gap (bps)",SUB); put(16,RQ+1,f"=($B${R_FIX}-$R$15)*100",OUT,nf=NUM)

# ================= SENSITIVITY / EFFICIENCY / CONE source
banner(18,RQ,RQ+NS+3,"NII × IRS NOTIONAL · efficiency & cone source")
put(SEN0,RQ,"p%",SUB)
for j in range(NS): put(SEN0,RQ+1+j,f"={gcl(8+j)}5",SUB,align="center")
put(SEN0,RQ+NS+1,"Disp.",SUB); put(SEN0,RQ+NS+2,"Exp(p)",SUB); put(SEN0,RQ+NS+3,"Vol(p)",SUB)
for i,p in enumerate(range(0,101,10)):
    r=SEN0+1+i; put(r,RQ,p,fill=CALC,nf="0")
    for j in range(NS):
        b=SB+j*BW; c=RQ+1+j
        put(r,c,f"={gcl(c)}$6+${gcl(RQ)}{r}/100*$B${R_CASH}*{gcl(b+7)}${U}-{gcl(c)}$8",fill=CALC,nf=NUM)
    rng=f"$R{r}:${gcl(RQ+NS)}{r}"
    for j in range(NS):  # hidden mirror: NA() when scenario weight is 0 (cols 60+)
        put(r,60+j,f"=IF({gcl(8+j)}${WRr}=0,NA(),{gcl(RQ+1+j)}{r})",fill=CALC)
    mr=f"${gcl(60)}{r}:${gcl(60+NS-1)}{r}"
    put(r,RQ+NS+1,f"=_xlfn.AGGREGATE(14,6,{mr},1)-_xlfn.AGGREGATE(15,6,{mr},1)",fill=CALC,nf=N2)
    put(r,RQ+NS+2,f"=SUMPRODUCT({wrng},{rng})/MAX(0.0001,{wsum})",fill=CALC,nf=NUM)
    put(r,RQ+NS+3,f"=SQRT(MAX(0,SUMPRODUCT({wrng},({rng}-{gcl(RQ+NS+2)}{r})^2)/MAX(0.0001,{wsum})))",fill=CALC,nf=N2)
dR=f"${gcl(RQ+NS+1)}${SEN0+1}:${gcl(RQ+NS+1)}${SEN0+11}"
put(31,RQ,"Min dispersion @ notional %",SUB)
put(31,RQ+2,f"=INDEX(${gcl(RQ)}${SEN0+1}:${gcl(RQ)}${SEN0+11},MATCH(MIN({dR}),{dR},0))",OUT,nf="0",
    note="The hedge ratio where weighted scenarios converge most - the cone's pinch.")
put(31,RQ+3,f"=MIN({dR})",OUT,nf=N2)

# ================= TIMELINE (selected scenario)
banner(4,TLc,TLc+8,"CASHFLOW COMPOSITION")
put(5,TLc,"Scenario:",SUB)
put(5,TLc+1,SCEN[2],BLUE,border=True,note="Drives the composition table, composition chart and NET chart.")
dv2=DataValidation(type="list",formula1=f"=$R$5:${gcl(RQ+NS)}$5"); ws.add_data_validation(dv2)
dv2.add(f"{gcl(TLc+1)}5")
SELC=f"${gcl(TLc+1)}$5"
mtch=f"MATCH({SELC},$R$5:${gcl(RQ+NS)}$5,0)-1"
for c,h in enumerate(["Month","ON","1M","2M","3M","SWAP","BOND","NET"]): put(6,TLc+c,h,SUB)
GE=gcl(SB+NS*BW)  # last col letter bound for INDEX row range
for m in range(NMX):
    r=7+m; gr=G0+m
    put(r,TLc,f"=A{gr}",fill=CALC,nf="MMM-YY")
    for k in range(4):
        put(r,TLc+1+k,
            f"=$B${AL0+k}/$B${R_CASH}*INDEX($A{gr}:${GE}{gr},{SB}+({mtch})*{BW}+5)"
            f"*INDEX($A{gr}:${GE}{gr},{SB}+({mtch})*{BW}+{k+1})/100*$B{gr}/360",fill=CALC,nf="0.000")
    put(r,TLc+5,f"=$B${R_NOT}/100*$B${R_CASH}*INDEX($A{gr}:${GE}{gr},{SB}+({mtch})*{BW}+7)",fill=CALC,nf="0.000")
    put(r,TLc+6,f"=-$C{gr}",fill=CALC,nf="0.000")
    put(r,TLc+7,f"=SUM({gcl(TLc+1)}{r}:{gcl(TLc+6)}{r})",OUT,nf="0.000")

# ================= SELF-TEST formulas
tests=[("Weights sum 100%",f'=IF(ABS(SUM(H{WRr}:{gcl(7+NS)}{WRr})-1)<0.001,"OK","FAIL")'),
("Allocation",'=IF(B22="OK","OK","FAIL")'),
("Tenor 6-36",f'=IF(AND($B${R_TEN}>=6,$B${R_TEN}<=36),"OK","FAIL")'),
("NII identity",f'=IF(SUMPRODUCT(ABS({nii}-($R$6:${gcl(RQ+NS)}$6+$R$7:${gcl(RQ+NS)}$7-$R$8:${gcl(RQ+NS)}$8)))<0.001,"OK","FAIL")'),
("Timeline=NII",f'=IF(ABS(SUM({gcl(TLc+7)}7:{gcl(TLc+7)}{6+NMX})-INDEX({nii},MATCH({SELC},$R$5:${gcl(RQ+NS)}$5,0)))<0.01,"OK","FAIL")'),
("Balance recon",f'=IF(ABS(({gcl(SB+5)}{G0+NMX-1}+{gcl(SB+6)}{G0+NMX-1}-$B${R_CASH})-{gcl(SB+6)}${U})<0.01,"OK","FAIL")')]
for i,(k,f) in enumerate(tests):
    r=T0r+1+i if i<5 else T0r+6
    put(T0r+1+i,1,k); c=put(T0r+1+i,2,f,SUB); c.fill=WARN
put(T1r+1,1,"MODEL STATUS",SUB)
c=put(T1r+1,2,f'=IF(COUNTIF(B{T0r+1}:B{T0r+6},"FAIL")=0,"OK — ALL TESTS PASS","CHECK FAILED TESTS")',
      Font(name=A,bold=True,color="006100")); c.fill=WARN
put(T1r+2,1,"Version",SUB); put(T1r+2,2,VERSION,SUB)

# ================= CHARTS dashboard
banner(CH0-2,1,24,"CHARTS")
def line(t,h=8.5,w=15):
    c=LineChart(); c.title=t; c.height,c.width=h,w
    c.y_axis.delete=False; c.x_axis.delete=False; return c
cats=Reference(ws,min_col=1,min_row=G0,max_row=G0+NMX-1)
ch=line("SOFR step paths (%)")
for j in range(NS):
    ch.add_data(Reference(ws,min_col=SB+j*BW,min_row=GH,max_row=G0+NMX-1),titles_from_data=True)
ch.set_categories(cats)
for sr in ch.series: sr.smooth=False
ws.add_chart(ch,f"A{CH0}")
# efficiency: scatter Vol(p) -> Exp(p)
ce=ScatterChart(); ce.title="Efficiency — expected NII vs volatility along hedge ratio"
ce.height,ce.width=8.5,15; ce.x_axis.title="Volatility"; ce.y_axis.title="Expected NII"
ce.x_axis.delete=False; ce.y_axis.delete=False
se=Series(Reference(ws,min_col=RQ+NS+2,min_row=SEN0+1,max_row=SEN0+11),
          Reference(ws,min_col=RQ+NS+3,min_row=SEN0+1,max_row=SEN0+11),title="IRS 0→100%")
se.marker=Marker(symbol="circle",size=6)
ce.series.append(se)
ws.add_chart(ce,f"J{CH0}")
ch2=line("Convergence cone — NII vs IRS notional %")
ch2.add_data(Reference(ws,min_col=RQ+1,max_col=RQ+NS,min_row=SEN0,max_row=SEN0+11),titles_from_data=True)
ch2.set_categories(Reference(ws,min_col=RQ,min_row=SEN0+1,max_row=SEN0+11))
for sr in ch2.series: sr.smooth=False
ws.add_chart(ch2,f"S{CH0}")
# expense: stacked area per bond
ca=AreaChart(); ca.grouping="stacked"; ca.title="Interest expense — composition by issue ($mm/month)"
ca.height,ca.width=8.5,15; ca.y_axis.delete=False; ca.x_axis.delete=False
ca.add_data(Reference(ws,min_col=4,max_col=3+NB,min_row=GH,max_row=G0+NMX-1),titles_from_data=True)
ca.set_categories(cats)
ws.add_chart(ca,f"A{CH0+19}")
ch5=BarChart(); ch5.type="col"; ch5.grouping="stacked"; ch5.overlap=100
ch5.title="Cashflow composition (selected scenario)"; ch5.height,ch5.width=8.5,15
ch5.y_axis.delete=False; ch5.x_axis.delete=False
ch5.add_data(Reference(ws,min_col=TLc+1,max_col=TLc+6,min_row=6,max_row=6+NMX),titles_from_data=True)
ch5.set_categories(Reference(ws,min_col=TLc,min_row=7,max_row=6+NMX))
ws.add_chart(ch5,f"J{CH0+19}")
ln=line("NET NII by month (selected scenario)")
ln.add_data(Reference(ws,min_col=TLc+7,min_row=6,max_row=6+NMX),titles_from_data=True)
ln.set_categories(Reference(ws,min_col=TLc,min_row=7,max_row=6+NMX))
for sr in ln.series: sr.smooth=False
ws.add_chart(ln,f"S{CH0+19}")

wb.calculation=CalcProperties(fullCalcOnLoad=True)
ws.sheet_properties.outlinePr.summaryBelow=True
ws.freeze_panes="A4"
for _c in range(60,60+NS): ws.column_dimensions[gcl(_c)].hidden=True
for col,w in [("A",26),("B",13),("C",9),("D",12),("F",12),("G",11),("Q",24)]:
    ws.column_dimensions[col].width=w
for j in range(NS): ws.column_dimensions[gcl(8+j)].width=11

import os
_dir=os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
_out=os.path.join(_dir,f"NII_Frontier_Lite_{VERSION}.xlsx")
try: wb.save(_out)
except PermissionError:
    _out=os.path.join(_dir,f"NII_Frontier_Lite_{VERSION}_new.xlsx"); wb.save(_out)
print(f"saved: {_out}")
