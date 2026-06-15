"""build_simple_example.py — creates NII_Engine_Simple.xlsx
A minimal workbook to test the NII engine. Ships formula-free; the user
imports the VBA (NII_ENGINE_ALL_VBA.txt) then pastes the formulas shown
in column E of the MODEL sheet. Run: python3 build_simple_example.py
Requires: openpyxl  (pip install openpyxl)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.properties import CalcProperties
from datetime import date

A="Arial"; NAVY="1F3864"
BLUE=Font(name=A,color="0000FF"); BLK=Font(name=A,size=10); SUB=Font(name=A,bold=True,size=10)
YEL=PatternFill("solid",start_color="FFF2CC"); MONO=Font(name="Consolas",size=9,color="7F4F00")
thin=Side(style="thin",color="BFBFBF"); BOX=Border(left=thin,right=thin,top=thin,bottom=thin)

def put(ws,r,c,v,font=BLK,fill=None,nf=None,box=False):
    x=ws.cell(row=r,column=c,value=v); x.font=font
    if fill:x.fill=fill
    if nf:x.number_format=nf
    if box:x.border=BOX
    return x

def build(path="NII_Engine_Simple.xlsx"):
    wb=Workbook()
    ws=wb.active; ws.title="README"; ws.sheet_view.showGridLines=False
    lines=["NII ENGINE — SIMPLE EXAMPLE","",
    "Minimal workbook to test the engine: build a calendar and a curve, then",
    "accrue interest and price a swap month. No formulas ship; you add the VBA,",
    "then paste the formulas from column E of the MODEL sheet.","",
    "SETUP",
    "1. Open NII_ENGINE_ALL_VBA.txt (6 modules, banner-separated).",
    "2. Alt+F11. Per module: Insert > Module (CLASS Module for cCalendar &",
    "   cStripCurve), rename, paste body, skip the 'Attribute VB_Name' line.",
    "3. Save as .xlsm. Press Ctrl+Alt+F9.",
    "4. MODEL: paste each column-E formula into the yellow cell on its left",
    "   (B4 calendar, B5 curve first).",
    "5. Results should match the Expected column.",
    "6. Alt+F8 > TestAll -> TESTS sheet, expect MODEL STATUS: OK.","",
    "RULE: handles flow cell-to-cell (B4, B5) — never type the handle as text."]
    for i,t in enumerate(lines,1):
        ws.cell(row=i,column=1,value=t).font=(Font(name=A,bold=True,size=15,color=NAVY) if i==1 else BLK)
    ws.column_dimensions["A"].width=92

    ws=wb.create_sheet("INPUTS"); ws.sheet_view.showGridLines=False
    put(ws,1,1,"INPUTS — objects rebuild from here on recalc",Font(name=A,bold=True,size=12,color=NAVY))
    put(ws,3,1,"HOLIDAYS (Excel Table — add rows freely)",SUB)
    put(ws,4,1,"Date",SUB); put(ws,4,2,"Name",SUB)
    hol=[(date(2026,1,1),"New Year"),(date(2026,1,19),"MLK Day"),(date(2026,7,3),"Independence (obs)"),
         (date(2026,9,7),"Labor Day"),(date(2026,11,26),"Thanksgiving"),(date(2026,12,25),"Christmas")]
    for i,(d,n) in enumerate(hol,5):
        put(ws,i,1,d,BLUE,nf="YYYY-MM-DD",box=True); put(ws,i,2,n,BLUE,box=True)
    t=Table(displayName="tblHolidays",ref="A4:B%d"%(4+len(hol)))
    t.tableStyleInfo=TableStyleInfo(name="TableStyleLight15",showRowStripes=True); ws.add_table(t)
    put(ws,3,4,"SCENARIO 'cuts' — FOMC moves",SUB)
    put(ws,4,4,"Decision",SUB); put(ws,4,5,"Move bps",SUB)
    for i,(d,m) in enumerate([(date(2026,6,17),-25),(date(2026,7,29),-25),(date(2026,9,16),-25),(date(2026,10,28),0)],5):
        put(ws,i,4,d,BLUE,nf="YYYY-MM-DD",box=True); put(ws,i,5,m,BLUE,nf='0;-0;"-"',box=True)
    put(ws,10,4,"Initial SOFR (%)",SUB); put(ws,10,5,3.80,BLUE,nf="0.00",box=True)
    put(ws,11,4,"Scale ×",SUB); put(ws,11,5,1.0,BLUE,nf="0.00",box=True)
    for c,w in [("A",13),("B",22),("D",14),("E",11)]: ws.column_dimensions[c].width=w

    ws=wb.create_sheet("MODEL"); ws.sheet_view.showGridLines=False
    put(ws,1,1,"MODEL — paste each column-E formula into the yellow cell",Font(name=A,bold=True,size=12,color=NAVY))
    put(ws,3,1,"1 · OBJECTS",SUB); put(ws,3,5,"FORMULA TO PASTE",SUB)
    put(ws,4,1,"Calendar handle"); put(ws,4,2,None,fill=YEL,box=True)
    put(ws,4,5,"=FED_HOL(tblHolidays[Date])",MONO)
    put(ws,5,1,"Curve handle"); put(ws,5,2,None,fill=YEL,box=True)
    put(ws,5,5,'=MAKE_CURVE("cuts",INPUTS!E10,INPUTS!D5:D8,INPUTS!E5:E8,INPUTS!E11,$B$4)',MONO)
    put(ws,7,1,"2 · ACCRUALS",SUB); put(ws,7,3,"Result",SUB); put(ws,7,4,"Expected",SUB); put(ws,7,5,"FORMULA TO PASTE",SUB)
    cases=[("ON, normal (15-Jul)",0.009861,'=ACCRUE(DATE(2026,7,15),DATE(2026,7,16),100,"SIMPLE",$B$5,"I")'),
    ("ON, weekend (17-Jul)",0.029583,'=ACCRUE(DATE(2026,7,17),DATE(2026,7,20),100,"SIMPLE",$B$5,"I")'),
    ("ON, holiday bridge (02-Jul)",0.039444,'=ACCRUE(DATE(2026,7,2),DATE(2026,7,6),100,"SIMPLE",$B$5,"I")'),
    ("Period simple (Jul-Oct)",0.853750,'=ACCRUE(DATE(2026,7,1),DATE(2026,10,1),100,"SIMPLE",$B$5,"I")'),
    ("Period compound (Jul-Oct)",0.857325,'=ACCRUE(DATE(2026,7,1),DATE(2026,10,1),100,"COMPOUND",$B$5,"I")'),
    ("Compound + principal",100.857325,'=ACCRUE(DATE(2026,7,1),DATE(2026,10,1),100,"COMPOUND",$B$5,"PI")')]
    r=8
    for nm,exp,f in cases:
        put(ws,r,1,nm); x=put(ws,r,2,None,fill=YEL,box=True); x.number_format="0.000000"
        put(ws,r,3,'=IF(ABS(B%d-D%d)<0.000001,"OK","≠")'%(r,r),SUB)
        put(ws,r,4,exp,nf="0.000000"); put(ws,r,5,f,MONO); r+=1
    put(ws,r+1,1,"3 · SWAP MONTH ($375mm @3.15%, Jul, RCV) — three legs",SUB); r+=2
    for nm,exp,f in [("FIXED leg",1.017187,'=SWAP_NET(DATE(2026,7,1),DATE(2026,8,1),375,3.15,$B$5,"RCV","FIXED")'),
    ("FLOAT leg",1.142773,'=SWAP_NET(DATE(2026,7,1),DATE(2026,8,1),375,3.15,$B$5,"RCV","FLOAT")'),
    ("NET",-0.125585,'=SWAP_NET(DATE(2026,7,1),DATE(2026,8,1),375,3.15,$B$5,"RCV","NET")')]:
        put(ws,r,1,nm); x=put(ws,r,2,None,fill=YEL,box=True); x.number_format="0.000000"
        put(ws,r,3,'=IF(ABS(B%d-D%d)<0.000001,"OK","≠")'%(r,r),SUB)
        put(ws,r,4,exp,nf="0.000000"); put(ws,r,5,f,MONO); r+=1
    put(ws,r+1,1,"NET = FIXED − FLOAT (RCV).",Font(name=A,size=9,color="595959"))
    for c,w in [("A",30),("B",13),("C",8),("D",13),("E",78)]: ws.column_dimensions[c].width=w

    wb.calculation=CalcProperties(fullCalcOnLoad=True)
    wb.save(path); print("saved:",path)

if __name__=="__main__":
    build()
